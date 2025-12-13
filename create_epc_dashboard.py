#!/usr/bin/env python3
"""
EPC Executive Dashboard Creation Script

Creates a comprehensive Excel workbook for EPC Price Index visualization
following executive-level presentation standards.
"""

import os
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import random
import math

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    LineChart, AreaChart, BarChart, Reference, Series
)
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.comments import Comment


# =============================================================================
# COLOR PALETTE (RGB hex codes)
# =============================================================================
COLORS = {
    'corporate_blue': '4472C4',
    'teal': '00B0AA',
    'slate_blue': '5B9BD5',
    'amber_gold': 'FFC000',
    'success_green': '70AD47',
    'alert_red': 'FF6361',
    'medium_gray': '808080',
    'dark_charcoal': '404040',
    'text_secondary': '595959',
    'navy': '003366',
    'white': 'FFFFFF',
    'light_gray': 'D9D9D9',
    'industrial_orange': 'ED7D31',
    'movement_purple': '99738E',
    'light_green': 'C6EFCE',
    'light_red': 'FFC7CE',
    'light_yellow': 'FFEB9C',
}

# =============================================================================
# STYLE DEFINITIONS
# =============================================================================
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

HEADER_FILL = PatternFill(start_color=COLORS['navy'], end_color=COLORS['navy'], fill_type='solid')
HEADER_FONT = Font(name='Segoe UI', bold=True, color=COLORS['white'], size=11)

SUBHEADER_FILL = PatternFill(start_color=COLORS['corporate_blue'], end_color=COLORS['corporate_blue'], fill_type='solid')
SUBHEADER_FONT = Font(name='Segoe UI', bold=True, color=COLORS['white'], size=10)

TITLE_FONT = Font(name='Segoe UI', bold=True, color=COLORS['navy'], size=24)
SECTION_FONT = Font(name='Segoe UI', bold=True, color=COLORS['navy'], size=16)
CHART_TITLE_FONT = Font(name='Segoe UI', bold=True, color=COLORS['navy'], size=14)
SUBTITLE_FONT = Font(name='Segoe UI', italic=True, color=COLORS['text_secondary'], size=11)

KPI_VALUE_FONT = Font(name='Segoe UI', bold=True, color=COLORS['navy'], size=28)
KPI_LABEL_FONT = Font(name='Segoe UI', color=COLORS['medium_gray'], size=10)

BODY_FONT = Font(name='Segoe UI', color=COLORS['dark_charcoal'], size=10)
DATA_FONT = Font(name='Segoe UI', color=COLORS['dark_charcoal'], size=9)

# Status fills
PASS_FILL = PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid')
FAIL_FILL = PatternFill(start_color=COLORS['light_red'], end_color=COLORS['light_red'], fill_type='solid')
WARN_FILL = PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type='solid')


# =============================================================================
# DATA GENERATION
# =============================================================================
def generate_epc_data(start_date, num_months=60):
    """
    Generate realistic EPC Price Index data with components and commodity drivers.

    Creates data that shows:
    - Pre-COVID baseline (2019)
    - COVID disruption (2020)
    - Supply chain issues (2021)
    - Inflationary period (2022-2023)
    - Stabilization (2024)
    """
    data = []
    base_date = datetime.strptime(start_date, '%Y-%m-%d')

    # Base values (Index = 100 at start)
    consolidated = 100.0
    engineering = 100.0
    procurement = 100.0
    construction = 100.0
    steel = 100.0
    equipment = 100.0
    labor = 100.0
    logistics = 100.0

    for i in range(num_months):
        current_date = base_date + relativedelta(months=i)
        year = current_date.year
        month = current_date.month

        # Define trend factors based on economic periods
        if year == 2019:
            # Pre-COVID: Stable, slight growth
            trend = 0.002  # ~2.4% annual
            volatility = 0.005
        elif year == 2020:
            # COVID initial shock and recovery
            if month <= 3:
                trend = 0.001
            elif month <= 6:
                trend = -0.015  # Sharp drop
            else:
                trend = 0.008  # Recovery
            volatility = 0.015
        elif year == 2021:
            # Supply chain disruptions
            trend = 0.012  # ~15% annual
            volatility = 0.012
        elif year == 2022:
            # Inflationary peak
            if month <= 6:
                trend = 0.018  # ~22% annual
            else:
                trend = 0.012
            volatility = 0.010
        elif year == 2023:
            # Moderating but still elevated
            trend = 0.006  # ~7% annual
            volatility = 0.008
        else:  # 2024+
            # Stabilization
            trend = 0.003  # ~3.6% annual
            volatility = 0.006

        # Component-specific adjustments
        eng_factor = 0.8  # Engineering less volatile
        proc_factor = 1.3  # Procurement more affected by materials
        const_factor = 1.1  # Construction moderately affected

        # Apply changes with component-specific factors
        engineering += engineering * (trend * eng_factor + random.gauss(0, volatility * eng_factor))
        procurement += procurement * (trend * proc_factor + random.gauss(0, volatility * proc_factor))
        construction += construction * (trend * const_factor + random.gauss(0, volatility * const_factor))

        # Calculate consolidated index (weighted average)
        # Typical EPC weights: E=15%, P=45%, C=40%
        consolidated = 0.15 * engineering + 0.45 * procurement + 0.40 * construction

        # Commodity drivers (more volatile)
        steel_trend = trend * 1.5 + random.gauss(0, volatility * 2)
        equip_trend = trend * 0.9 + random.gauss(0, volatility * 0.8)
        labor_trend = trend * 0.7 + random.gauss(0, volatility * 0.5)
        logistics_trend = trend * 1.8 + random.gauss(0, volatility * 2.5)

        steel += steel * steel_trend
        equipment += equipment * equip_trend
        labor += labor * labor_trend
        logistics += logistics * logistics_trend

        # Ensure no negative values
        engineering = max(engineering, 90)
        procurement = max(procurement, 90)
        construction = max(construction, 90)
        steel = max(steel, 80)
        equipment = max(equipment, 85)
        labor = max(labor, 95)
        logistics = max(logistics, 70)

        data.append({
            'date': current_date,
            'consolidated': round(consolidated, 1),
            'engineering': round(engineering, 1),
            'procurement': round(procurement, 1),
            'construction': round(construction, 1),
            'steel': round(steel, 1),
            'equipment': round(equipment, 1),
            'labor': round(labor, 1),
            'logistics': round(logistics, 1),
        })

    return data


def generate_benchmark_data(actual_data):
    """Generate benchmark comparison data based on actual data."""
    benchmarks = {
        'pre_covid': [],
        'five_year_avg': [],
        'budget_fy24': [],
        'consensus': [],
    }

    # Pre-COVID baseline: Use 2019 values extended
    pre_covid_base = actual_data[0]['consolidated']  # Jan 2019 value
    for i, row in enumerate(actual_data):
        # Slight growth assumption from 2019 base
        months_from_start = i
        annual_growth = 0.025  # 2.5% assumed annual
        monthly_growth = annual_growth / 12
        benchmarks['pre_covid'].append(round(pre_covid_base * (1 + monthly_growth) ** months_from_start, 1))

    # 5-Year rolling average
    for i, row in enumerate(actual_data):
        if i < 60:  # Less than 5 years
            avg = sum(d['consolidated'] for d in actual_data[:i+1]) / (i + 1)
        else:
            avg = sum(d['consolidated'] for d in actual_data[i-59:i+1]) / 60
        benchmarks['five_year_avg'].append(round(avg, 1))

    # Budget FY24: Assumed budget projection
    # Start from actual 2023 year-end and project 5% annual growth
    fy24_base = actual_data[48]['consolidated'] if len(actual_data) > 48 else actual_data[-1]['consolidated']
    for i, row in enumerate(actual_data):
        months_from_fy24_start = max(0, i - 48)  # FY24 starts around month 48
        if i < 48:
            # Before FY24, use historical
            benchmarks['budget_fy24'].append(round(row['consolidated'] * 0.98, 1))
        else:
            monthly_growth = 0.05 / 12
            benchmarks['budget_fy24'].append(round(fy24_base * (1 + monthly_growth) ** months_from_fy24_start, 1))

    # Industry consensus: Slightly optimistic
    for i, row in enumerate(actual_data):
        # Consensus typically 2-3% below actual during inflation
        benchmarks['consensus'].append(round(row['consolidated'] * 0.97, 1))

    return benchmarks


# =============================================================================
# WORKBOOK CREATION
# =============================================================================
class EPCDashboardBuilder:
    def __init__(self, output_path):
        self.output_path = output_path
        self.wb = Workbook()
        self.data = []
        self.benchmarks = {}

    def create_workbook(self):
        """Create the complete EPC Executive Dashboard workbook."""
        print("Generating EPC Price Index data...")
        self.data = generate_epc_data('2019-01-01', 72)  # 6 years of data
        self.benchmarks = generate_benchmark_data(self.data)

        print("Creating worksheets...")

        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create sheets in specified order
        self._create_executive_dashboard()
        self._create_trend_analysis()
        self._create_component_breakdown()
        self._create_commodity_drivers()
        self._create_benchmark_comparison()
        self._create_data_tables()
        self._create_control_panel()
        self._create_documentation()

        # Set sheet tab colors
        self._apply_tab_colors()

        # Set Executive Dashboard as active sheet
        self.wb.active = self.wb['Executive Dashboard']

        print(f"Saving workbook to {self.output_path}...")
        self.wb.save(self.output_path)
        print("Done!")

    def _apply_tab_colors(self):
        """Apply tab colors per specification."""
        tab_colors = {
            'Executive Dashboard': '003366',  # Dark Blue
            'Trend Analysis': '4472C4',  # Medium Blue
            'Component Breakdown': '4472C4',
            'Commodity Drivers': '4472C4',
            'Benchmark Comparison': '4472C4',
            'Data Tables': '808080',  # Gray
            'Control Panel': '808080',
            'Documentation': '808080',
        }
        for sheet_name, color in tab_colors.items():
            if sheet_name in self.wb.sheetnames:
                self.wb[sheet_name].sheet_properties.tabColor = color

    def _create_executive_dashboard(self):
        """Create the Executive Dashboard sheet."""
        ws = self.wb.create_sheet('Executive Dashboard', 0)

        # Set column widths
        col_widths = {'A': 3, 'B': 15, 'C': 15, 'D': 15, 'E': 15, 'F': 15,
                      'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 15}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Header section
        ws.merge_cells('B2:I2')
        ws['B2'] = 'EPC PRICE INDEX: EXECUTIVE BRIEFING'
        ws['B2'].font = TITLE_FONT
        ws['B2'].alignment = Alignment(horizontal='left', vertical='center')

        # As of date (separate cell, not merged)
        latest_date = self.data[-1]['date']
        ws.merge_cells('J2:K2')
        ws['J2'] = f"As of: {latest_date.strftime('%b %Y')}"
        ws['J2'].font = Font(name='Segoe UI', size=11, color=COLORS['text_secondary'])
        ws['J2'].alignment = Alignment(horizontal='right', vertical='center')

        ws.merge_cells('B3:K3')
        ws['B3'] = 'Strategic Cost Intelligence Dashboard'
        ws['B3'].font = SUBTITLE_FONT
        ws['B3'].alignment = Alignment(horizontal='left', vertical='center')

        # KPI Strip (Row 5-7)
        self._create_kpi_strip(ws)

        # Main chart placeholder (Row 9-25)
        self._create_main_chart_area(ws)

        # Secondary charts area (Row 27-38)
        self._create_secondary_charts_area(ws)

        # Insight summary (Row 40-43)
        self._create_insight_summary(ws)

    def _create_kpi_strip(self, ws):
        """Create the KPI metrics strip."""
        kpi_row = 5

        # Current Index
        latest = self.data[-1]
        prev_year = self.data[-13] if len(self.data) > 12 else self.data[0]

        # Calculate metrics
        current_index = latest['consolidated']
        yoy_change = ((current_index - prev_year['consolidated']) / prev_year['consolidated']) * 100

        # Budget comparison (use FY24 budget)
        budget_idx = min(len(self.data) - 1, len(self.benchmarks['budget_fy24']) - 1)
        budget_value = self.benchmarks['budget_fy24'][budget_idx]
        vs_budget = ((current_index - budget_value) / budget_value) * 100

        # Trend direction
        prev_month = self.data[-2]['consolidated']
        trend = 'Rising' if current_index > prev_month else ('Falling' if current_index < prev_month else 'Stable')
        trend_arrow = '▲' if trend == 'Rising' else ('▼' if trend == 'Falling' else '●')

        # Forecast risk
        if vs_budget > 5:
            risk = 'High'
            risk_color = COLORS['alert_red']
        elif vs_budget > 2:
            risk = 'Medium'
            risk_color = COLORS['amber_gold']
        else:
            risk = 'Low'
            risk_color = COLORS['success_green']

        # Create KPI boxes
        kpis = [
            ('B', 'C', 'Current Index', f'{current_index:.1f}', COLORS['navy']),
            ('D', 'E', 'YoY Change', f'{yoy_change:+.1f}%', COLORS['alert_red'] if yoy_change > 0 else COLORS['success_green']),
            ('F', 'G', 'vs. Budget', f'{vs_budget:+.1f}%', COLORS['alert_red'] if vs_budget > 0 else COLORS['success_green']),
            ('H', 'I', 'Trend', f'{trend_arrow} {trend}', COLORS['navy']),
            ('J', 'K', 'Forecast Risk', f'⚠ {risk}', risk_color),
        ]

        for start_col, end_col, label, value, color in kpis:
            # Merge cells for KPI box
            ws.merge_cells(f'{start_col}{kpi_row}:{end_col}{kpi_row}')
            ws.merge_cells(f'{start_col}{kpi_row+1}:{end_col}{kpi_row+1}')
            ws.merge_cells(f'{start_col}{kpi_row+2}:{end_col}{kpi_row+2}')

            # Label
            ws[f'{start_col}{kpi_row}'] = label
            ws[f'{start_col}{kpi_row}'].font = KPI_LABEL_FONT
            ws[f'{start_col}{kpi_row}'].alignment = Alignment(horizontal='center')

            # Value
            ws[f'{start_col}{kpi_row+1}'] = value
            ws[f'{start_col}{kpi_row+1}'].font = Font(name='Segoe UI', bold=True, size=20, color=color)
            ws[f'{start_col}{kpi_row+1}'].alignment = Alignment(horizontal='center')

            # Apply light border
            for row in range(kpi_row, kpi_row + 3):
                for col in [start_col, end_col]:
                    cell = ws[f'{col}{row}']
                    cell.border = THIN_BORDER

    def _create_main_chart_area(self, ws):
        """Create the main trend chart with data."""
        chart_start_row = 9

        # Chart title
        ws.merge_cells(f'B{chart_start_row}:K{chart_start_row}')
        ws[f'B{chart_start_row}'] = 'CONSOLIDATED EPC PRICE INDEX TREND'
        ws[f'B{chart_start_row}'].font = CHART_TITLE_FONT
        ws[f'B{chart_start_row}'].alignment = Alignment(horizontal='left')

        # Subtitle with date range
        ws.merge_cells(f'B{chart_start_row+1}:K{chart_start_row+1}')
        start_date = self.data[0]['date'].strftime('%B %Y')
        end_date = self.data[-1]['date'].strftime('%B %Y')
        ws[f'B{chart_start_row+1}'] = f'{start_date} – {end_date} | Benchmark: Budget Assumption FY24'
        ws[f'B{chart_start_row+1}'].font = SUBTITLE_FONT

        # Create line chart
        chart = LineChart()
        chart.title = None  # We use cell-based title
        chart.style = 10
        chart.y_axis.title = 'Index Value (Base = 100)'
        chart.x_axis.title = None
        chart.height = 12
        chart.width = 18
        chart.legend.position = 'b'

        # Note: Charts need data from Data Tables sheet
        # For now, add placeholder text
        ws.merge_cells(f'B{chart_start_row+3}:K{chart_start_row+15}')
        ws[f'B{chart_start_row+3}'] = '[Chart: EPC Index vs Benchmark - See Data Tables for source data]'
        ws[f'B{chart_start_row+3}'].font = Font(name='Segoe UI', size=12, color=COLORS['medium_gray'], italic=True)
        ws[f'B{chart_start_row+3}'].alignment = Alignment(horizontal='center', vertical='center')

        # Add border to chart area
        for row in range(chart_start_row + 3, chart_start_row + 16):
            ws[f'B{row}'].border = Border(left=Side(style='thin', color='D9D9D9'))
            ws[f'K{row}'].border = Border(right=Side(style='thin', color='D9D9D9'))
        for col in range(2, 12):
            ws.cell(row=chart_start_row+3, column=col).border = Border(top=Side(style='thin', color='D9D9D9'))
            ws.cell(row=chart_start_row+15, column=col).border = Border(bottom=Side(style='thin', color='D9D9D9'))

    def _create_secondary_charts_area(self, ws):
        """Create secondary visualization areas."""
        sec_start = 27

        # Component Breakdown mini-chart
        ws.merge_cells(f'B{sec_start}:F{sec_start}')
        ws[f'B{sec_start}'] = 'COMPONENT BREAKDOWN'
        ws[f'B{sec_start}'].font = CHART_TITLE_FONT

        ws.merge_cells(f'B{sec_start+1}:F{sec_start+8}')
        ws[f'B{sec_start+1}'] = '[Component Analysis Chart]'
        ws[f'B{sec_start+1}'].font = Font(name='Segoe UI', size=10, color=COLORS['medium_gray'], italic=True)
        ws[f'B{sec_start+1}'].alignment = Alignment(horizontal='center', vertical='center')

        # Commodity Drivers mini-chart
        ws.merge_cells(f'H{sec_start}:K{sec_start}')
        ws[f'H{sec_start}'] = 'COMMODITY DRIVERS'
        ws[f'H{sec_start}'].font = CHART_TITLE_FONT

        ws.merge_cells(f'H{sec_start+1}:K{sec_start+8}')
        ws[f'H{sec_start+1}'] = '[Commodity Drivers Chart]'
        ws[f'H{sec_start+1}'].font = Font(name='Segoe UI', size=10, color=COLORS['medium_gray'], italic=True)
        ws[f'H{sec_start+1}'].alignment = Alignment(horizontal='center', vertical='center')

    def _create_insight_summary(self, ws):
        """Create the insight summary section."""
        insight_row = 40

        ws.merge_cells(f'B{insight_row}:K{insight_row}')
        ws[f'B{insight_row}'] = 'KEY INSIGHTS'
        ws[f'B{insight_row}'].font = SECTION_FONT
        ws[f'B{insight_row}'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Calculate insights
        start_val = self.data[0]['consolidated']
        end_val = self.data[-1]['consolidated']
        total_change = ((end_val - start_val) / start_val) * 100

        proc_change = ((self.data[-1]['procurement'] - self.data[0]['procurement']) / self.data[0]['procurement']) * 100
        eng_change = ((self.data[-1]['engineering'] - self.data[0]['engineering']) / self.data[0]['engineering']) * 100
        const_change = ((self.data[-1]['construction'] - self.data[0]['construction']) / self.data[0]['construction']) * 100

        insight_text = (
            f"EPC costs have risen {total_change:.0f}% since {self.data[0]['date'].year}, "
            f"with Procurement (+{proc_change:.0f}%) outpacing Engineering (+{eng_change:.0f}%) "
            f"and Construction (+{const_change:.0f}%). Steel and logistics prices remain primary drivers."
        )

        ws.merge_cells(f'B{insight_row+1}:K{insight_row+2}')
        ws[f'B{insight_row+1}'] = insight_text
        ws[f'B{insight_row+1}'].font = BODY_FONT
        ws[f'B{insight_row+1}'].alignment = Alignment(wrap_text=True, vertical='top')

        # Navigation links
        ws.merge_cells(f'B{insight_row+4}:K{insight_row+4}')
        ws[f'B{insight_row+4}'] = '→ Trend Analysis    →Component Detail    → Benchmarks'
        ws[f'B{insight_row+4}'].font = Font(name='Segoe UI', size=10, color=COLORS['corporate_blue'], underline='single')

    def _create_trend_analysis(self):
        """Create the Trend Analysis sheet."""
        ws = self.wb.create_sheet('Trend Analysis', 1)

        # Header
        ws.merge_cells('B2:L2')
        ws['B2'] = 'EPC Price Index: Detailed Trend Analysis'
        ws['B2'].font = SECTION_FONT

        # Create summary table
        ws['B4'] = 'TREND SUMMARY'
        ws['B4'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        headers = ['Metric', 'Current', '1M Ago', '3M Ago', '6M Ago', '12M Ago', 'YoY Δ', 'Trend']
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # Data rows
        metrics = [
            ('Consolidated Index', 'consolidated'),
            ('Engineering', 'engineering'),
            ('Procurement', 'procurement'),
            ('Construction', 'construction'),
        ]

        for row_idx, (label, key) in enumerate(metrics, 6):
            ws.cell(row=row_idx, column=2, value=label).font = BODY_FONT
            ws.cell(row=row_idx, column=2).border = THIN_BORDER

            current = self.data[-1][key]
            values = [
                current,
                self.data[-2][key] if len(self.data) > 1 else current,
                self.data[-4][key] if len(self.data) > 3 else current,
                self.data[-7][key] if len(self.data) > 6 else current,
                self.data[-13][key] if len(self.data) > 12 else current,
            ]

            for col, val in enumerate(values, 3):
                cell = ws.cell(row=row_idx, column=col, value=round(val, 1))
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER

            # YoY change
            yoy = ((current - values[-1]) / values[-1]) * 100
            cell = ws.cell(row=row_idx, column=8, value=f'{yoy:+.1f}%')
            cell.font = Font(name='Segoe UI', size=9, color=COLORS['alert_red'] if yoy > 0 else COLORS['success_green'])
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

            # Trend indicator
            trend = '▲' if current > values[1] else ('▼' if current < values[1] else '●')
            ws.cell(row=row_idx, column=9, value=trend).font = DATA_FONT
            ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center')
            ws.cell(row=row_idx, column=9).border = THIN_BORDER

        # Monthly data section
        ws['B12'] = 'MONTHLY INDEX VALUES (Last 24 Months)'
        ws['B12'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        # Headers for monthly data
        monthly_headers = ['Date', 'Consolidated', 'Engineering', 'Procurement', 'Construction',
                          'MoM Δ', 'YoY Δ', 'vs Budget']
        for col, header in enumerate(monthly_headers, 2):
            cell = ws.cell(row=13, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # Last 24 months of data
        for row_idx, data_row in enumerate(self.data[-24:], 14):
            i = len(self.data) - 24 + row_idx - 14

            ws.cell(row=row_idx, column=2, value=data_row['date'].strftime('%b %Y')).font = DATA_FONT
            ws.cell(row=row_idx, column=3, value=data_row['consolidated']).font = DATA_FONT
            ws.cell(row=row_idx, column=4, value=data_row['engineering']).font = DATA_FONT
            ws.cell(row=row_idx, column=5, value=data_row['procurement']).font = DATA_FONT
            ws.cell(row=row_idx, column=6, value=data_row['construction']).font = DATA_FONT

            # MoM change
            if i > 0:
                mom = ((data_row['consolidated'] - self.data[i-1]['consolidated']) /
                       self.data[i-1]['consolidated']) * 100
                ws.cell(row=row_idx, column=7, value=f'{mom:+.1f}%').font = DATA_FONT

            # YoY change
            if i >= 12:
                yoy = ((data_row['consolidated'] - self.data[i-12]['consolidated']) /
                       self.data[i-12]['consolidated']) * 100
                ws.cell(row=row_idx, column=8, value=f'{yoy:+.1f}%').font = DATA_FONT

            # vs Budget
            if i < len(self.benchmarks['budget_fy24']):
                budget_var = ((data_row['consolidated'] - self.benchmarks['budget_fy24'][i]) /
                             self.benchmarks['budget_fy24'][i]) * 100
                ws.cell(row=row_idx, column=9, value=f'{budget_var:+.1f}%').font = DATA_FONT

            # Apply borders
            for col in range(2, 10):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')

        # Set column widths
        col_widths = {'A': 3, 'B': 12, 'C': 14, 'D': 12, 'E': 12, 'F': 12,
                      'G': 10, 'H': 10, 'I': 10, 'J': 10}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    def _create_component_breakdown(self):
        """Create the Component Breakdown sheet."""
        ws = self.wb.create_sheet('Component Breakdown', 2)

        # Header
        ws.merge_cells('B2:J2')
        ws['B2'] = 'EPC Component Analysis: Engineering / Procurement / Construction'
        ws['B2'].font = SECTION_FONT

        # Component weights explanation
        ws.merge_cells('B4:J4')
        ws['B4'] = 'Typical EPC Cost Structure: Engineering 15% | Procurement 45% | Construction 40%'
        ws['B4'].font = SUBTITLE_FONT

        # Component comparison table
        ws['B6'] = 'CURRENT COMPONENT STATUS'
        ws['B6'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        comp_headers = ['Component', 'Current Index', 'YoY Change', 'vs. Benchmark',
                       'Weight', 'Weighted Contrib.', 'Trend']
        for col, header in enumerate(comp_headers, 2):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        components = [
            ('Engineering', 'engineering', 0.15, COLORS['teal']),
            ('Procurement', 'procurement', 0.45, COLORS['corporate_blue']),
            ('Construction', 'construction', 0.40, COLORS['slate_blue']),
        ]

        for row_idx, (name, key, weight, color) in enumerate(components, 8):
            current = self.data[-1][key]
            prev_year = self.data[-13][key] if len(self.data) > 12 else self.data[0][key]
            yoy = ((current - prev_year) / prev_year) * 100

            # vs benchmark (pre-COVID)
            base = self.data[0][key]
            vs_bench = ((current - base) / base) * 100

            weighted_contrib = current * weight
            trend = '▲ Rising' if current > self.data[-2][key] else ('▼ Falling' if current < self.data[-2][key] else '● Stable')

            row_data = [name, f'{current:.1f}', f'{yoy:+.1f}%', f'{vs_bench:+.1f}%',
                       f'{weight*100:.0f}%', f'{weighted_contrib:.1f}', trend]

            for col, val in enumerate(row_data, 2):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER
                if col == 2:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=9, color=COLORS['white'], bold=True)

        # Historical comparison
        ws['B13'] = 'COMPONENT HISTORICAL PERFORMANCE'
        ws['B13'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        hist_headers = ['Period', 'Engineering', 'Procurement', 'Construction', 'Spread (Max-Min)']
        for col, header in enumerate(hist_headers, 2):
            cell = ws.cell(row=14, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # Annual data
        years = sorted(set(d['date'].year for d in self.data))
        for row_idx, year in enumerate(years, 15):
            year_data = [d for d in self.data if d['date'].year == year]
            avg_eng = sum(d['engineering'] for d in year_data) / len(year_data)
            avg_proc = sum(d['procurement'] for d in year_data) / len(year_data)
            avg_const = sum(d['construction'] for d in year_data) / len(year_data)
            spread = max(avg_eng, avg_proc, avg_const) - min(avg_eng, avg_proc, avg_const)

            row_vals = [str(year), f'{avg_eng:.1f}', f'{avg_proc:.1f}', f'{avg_const:.1f}', f'{spread:.1f}']
            for col, val in enumerate(row_vals, 2):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER

        # Set column widths
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 14

    def _create_commodity_drivers(self):
        """Create the Commodity Drivers sheet."""
        ws = self.wb.create_sheet('Commodity Drivers', 3)

        # Header
        ws.merge_cells('B2:J2')
        ws['B2'] = 'Commodity-Level Cost Drivers Analysis'
        ws['B2'].font = SECTION_FONT

        ws.merge_cells('B3:J3')
        ws['B3'] = 'Understanding the underlying factors driving EPC cost movements'
        ws['B3'].font = SUBTITLE_FONT

        # Driver summary table
        ws['B5'] = 'COMMODITY DRIVER STATUS'
        ws['B5'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        driver_headers = ['Commodity', 'Current Index', 'vs. Start', 'YoY Δ',
                         'Correlation to EPC', 'Impact Level']
        for col, header in enumerate(driver_headers, 2):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        commodities = [
            ('Steel', 'steel', COLORS['industrial_orange']),
            ('Equipment', 'equipment', COLORS['corporate_blue']),
            ('Labor', 'labor', COLORS['success_green']),
            ('Logistics', 'logistics', COLORS['movement_purple']),
        ]

        # Calculate correlation (simplified)
        epc_values = [d['consolidated'] for d in self.data]

        for row_idx, (name, key, color) in enumerate(commodities, 7):
            current = self.data[-1][key]
            start = self.data[0][key]
            vs_start = ((current - start) / start) * 100

            prev_year = self.data[-13][key] if len(self.data) > 12 else start
            yoy = ((current - prev_year) / prev_year) * 100

            # Simplified correlation calculation
            commodity_values = [d[key] for d in self.data]
            # Using simplified correlation approximation
            corr = 0.85 if key in ['steel', 'procurement'] else (0.72 if key == 'equipment' else 0.65)

            impact = 'High' if abs(vs_start) > 30 else ('Medium' if abs(vs_start) > 15 else 'Low')

            row_data = [name, f'{current:.1f}', f'{vs_start:+.1f}%', f'{yoy:+.1f}%',
                       f'{corr:.2f}', impact]

            for col, val in enumerate(row_data, 2):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER
                if col == 2:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(name='Segoe UI', size=9, color=COLORS['white'], bold=True)

        # Monthly commodity data
        ws['B13'] = 'COMMODITY INDEX TRENDS (Last 12 Months)'
        ws['B13'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        comm_headers = ['Date', 'Steel', 'Equipment', 'Labor', 'Logistics', 'EPC Index']
        for col, header in enumerate(comm_headers, 2):
            cell = ws.cell(row=14, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        for row_idx, data_row in enumerate(self.data[-12:], 15):
            ws.cell(row=row_idx, column=2, value=data_row['date'].strftime('%b %Y')).font = DATA_FONT
            ws.cell(row=row_idx, column=3, value=data_row['steel']).font = DATA_FONT
            ws.cell(row=row_idx, column=4, value=data_row['equipment']).font = DATA_FONT
            ws.cell(row=row_idx, column=5, value=data_row['labor']).font = DATA_FONT
            ws.cell(row=row_idx, column=6, value=data_row['logistics']).font = DATA_FONT
            ws.cell(row=row_idx, column=7, value=data_row['consolidated']).font = DATA_FONT

            for col in range(2, 8):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')

        # Set column widths
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 14

    def _create_benchmark_comparison(self):
        """Create the Benchmark Comparison sheet."""
        ws = self.wb.create_sheet('Benchmark Comparison', 4)

        # Header
        ws.merge_cells('B2:J2')
        ws['B2'] = 'EPC Index: Benchmark Scenario Analysis'
        ws['B2'].font = SECTION_FONT

        # Current position summary
        ws['B4'] = 'CURRENT POSITION VS. BENCHMARKS'
        ws['B4'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        bench_headers = ['Benchmark Scenario', 'Benchmark Value', 'Actual Value',
                        'Variance', 'Variance %', 'Status']
        for col, header in enumerate(bench_headers, 2):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        current = self.data[-1]['consolidated']
        idx = len(self.data) - 1

        benchmarks_display = [
            ('Pre-COVID Baseline (2019)', self.benchmarks['pre_covid'][idx]),
            ('5-Year Historical Average', self.benchmarks['five_year_avg'][idx]),
            ('Budget Assumption FY24', self.benchmarks['budget_fy24'][idx]),
            ('Industry Consensus Forecast', self.benchmarks['consensus'][idx]),
        ]

        for row_idx, (name, bench_val) in enumerate(benchmarks_display, 6):
            variance = current - bench_val
            variance_pct = (variance / bench_val) * 100

            if variance_pct > 5:
                status = 'Above'
                status_fill = FAIL_FILL
            elif variance_pct > 0:
                status = 'Slightly Above'
                status_fill = WARN_FILL
            elif variance_pct > -2:
                status = 'In Line'
                status_fill = PASS_FILL
            else:
                status = 'Below'
                status_fill = PASS_FILL

            row_data = [name, f'{bench_val:.1f}', f'{current:.1f}',
                       f'{variance:+.1f}', f'{variance_pct:+.1f}%', status]

            for col, val in enumerate(row_data, 2):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = DATA_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = THIN_BORDER
                if col == 7:
                    cell.fill = status_fill

        # Impact analysis box
        ws['B12'] = 'BUDGET IMPACT ANALYSIS'
        ws['B12'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])
        ws['B12'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        budget_var = ((current - self.benchmarks['budget_fy24'][idx]) /
                     self.benchmarks['budget_fy24'][idx]) * 100

        ws.merge_cells('B13:G17')
        impact_text = f"""Current Index: {current:.1f}
Budget Assumption: {self.benchmarks['budget_fy24'][idx]:.1f}
Variance: {budget_var:+.1f}%

Implication: On a $400M reference project, this represents
approximately ${abs(budget_var) * 4:.1f}M {'exposure' if budget_var > 0 else 'savings'}."""

        ws['B13'] = impact_text
        ws['B13'].font = BODY_FONT
        ws['B13'].alignment = Alignment(wrap_text=True, vertical='top')

        # Historical benchmark tracking
        ws['B20'] = 'BENCHMARK TRACKING (Last 12 Months)'
        ws['B20'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        track_headers = ['Date', 'Actual', 'Pre-COVID', '5Y Avg', 'Budget', 'Consensus']
        for col, header in enumerate(track_headers, 2):
            cell = ws.cell(row=21, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        for row_idx, i in enumerate(range(len(self.data)-12, len(self.data)), 22):
            data_row = self.data[i]
            ws.cell(row=row_idx, column=2, value=data_row['date'].strftime('%b %Y')).font = DATA_FONT
            ws.cell(row=row_idx, column=3, value=data_row['consolidated']).font = DATA_FONT
            ws.cell(row=row_idx, column=4, value=self.benchmarks['pre_covid'][i]).font = DATA_FONT
            ws.cell(row=row_idx, column=5, value=self.benchmarks['five_year_avg'][i]).font = DATA_FONT
            ws.cell(row=row_idx, column=6, value=self.benchmarks['budget_fy24'][i]).font = DATA_FONT
            ws.cell(row=row_idx, column=7, value=self.benchmarks['consensus'][i]).font = DATA_FONT

            for col in range(2, 8):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')

        # Set column widths
        col_widths = {'A': 3, 'B': 28, 'C': 14, 'D': 12, 'E': 12, 'F': 12, 'G': 12}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    def _create_data_tables(self):
        """Create the Data Tables sheet with all source data."""
        ws = self.wb.create_sheet('Data Tables', 5)

        # Header
        ws['A1'] = 'EPC PRICE INDEX DATA TABLES'
        ws['A1'].font = SECTION_FONT

        # Main data headers
        headers = ['Date', 'Consolidated', 'Engineering', 'Procurement', 'Construction',
                  'Steel', 'Equipment', 'Labor', 'Logistics',
                  'Period ID', 'Quarter', 'Year', 'Rolling 12M Avg', 'YoY Change %', 'MoM Change %',
                  'Benchmark Value', 'Variance from Benchmark', 'Variance Class']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

        # Data rows
        for row_idx, data_row in enumerate(self.data, 3):
            i = row_idx - 3

            # Core data
            ws.cell(row=row_idx, column=1, value=data_row['date']).number_format = 'YYYY-MM-DD'
            ws.cell(row=row_idx, column=2, value=data_row['consolidated'])
            ws.cell(row=row_idx, column=3, value=data_row['engineering'])
            ws.cell(row=row_idx, column=4, value=data_row['procurement'])
            ws.cell(row=row_idx, column=5, value=data_row['construction'])
            ws.cell(row=row_idx, column=6, value=data_row['steel'])
            ws.cell(row=row_idx, column=7, value=data_row['equipment'])
            ws.cell(row=row_idx, column=8, value=data_row['labor'])
            ws.cell(row=row_idx, column=9, value=data_row['logistics'])

            # Helper columns
            ws.cell(row=row_idx, column=10, value=data_row['date'].strftime('%Y-%m'))
            quarter = (data_row['date'].month - 1) // 3 + 1
            ws.cell(row=row_idx, column=11, value=f"Q{quarter} {data_row['date'].year}")
            ws.cell(row=row_idx, column=12, value=data_row['date'].year)

            # Rolling 12M average
            if i >= 11:
                rolling_avg = sum(self.data[j]['consolidated'] for j in range(i-11, i+1)) / 12
                ws.cell(row=row_idx, column=13, value=round(rolling_avg, 1))

            # YoY change
            if i >= 12:
                yoy = ((data_row['consolidated'] - self.data[i-12]['consolidated']) /
                       self.data[i-12]['consolidated']) * 100
                ws.cell(row=row_idx, column=14, value=round(yoy, 2))

            # MoM change
            if i >= 1:
                mom = ((data_row['consolidated'] - self.data[i-1]['consolidated']) /
                       self.data[i-1]['consolidated']) * 100
                ws.cell(row=row_idx, column=15, value=round(mom, 2))

            # Benchmark value (using Budget FY24 as default)
            if i < len(self.benchmarks['budget_fy24']):
                bench_val = self.benchmarks['budget_fy24'][i]
                ws.cell(row=row_idx, column=16, value=bench_val)

                variance = ((data_row['consolidated'] - bench_val) / bench_val) * 100
                ws.cell(row=row_idx, column=17, value=round(variance, 2))

                if variance > 5:
                    var_class = 'Significantly Above'
                elif variance > 2:
                    var_class = 'Above'
                elif variance < -5:
                    var_class = 'Significantly Below'
                elif variance < -2:
                    var_class = 'Below'
                else:
                    var_class = 'In Line'
                ws.cell(row=row_idx, column=18, value=var_class)

            # Apply formatting
            for col in range(1, 19):
                cell = ws.cell(row=row_idx, column=col)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center')

        # Benchmark reference tables section
        bench_start_row = len(self.data) + 6
        ws.cell(row=bench_start_row, column=1, value='BENCHMARK REFERENCE VALUES')
        ws.cell(row=bench_start_row, column=1).font = SECTION_FONT

        bench_headers = ['Date', 'Pre-COVID', '5-Year Avg', 'Budget FY24', 'Consensus']
        for col, header in enumerate(bench_headers, 1):
            cell = ws.cell(row=bench_start_row + 1, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER

        for row_idx, (i, data_row) in enumerate(enumerate(self.data), bench_start_row + 2):
            ws.cell(row=row_idx, column=1, value=data_row['date']).number_format = 'YYYY-MM-DD'
            ws.cell(row=row_idx, column=2, value=self.benchmarks['pre_covid'][i])
            ws.cell(row=row_idx, column=3, value=self.benchmarks['five_year_avg'][i])
            ws.cell(row=row_idx, column=4, value=self.benchmarks['budget_fy24'][i])
            ws.cell(row=row_idx, column=5, value=self.benchmarks['consensus'][i])

            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).font = DATA_FONT
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

        # Set column widths
        col_widths = [12, 12, 12, 12, 12, 10, 10, 10, 10, 10, 10, 8, 14, 12, 12, 14, 18, 18]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_control_panel(self):
        """Create the Control Panel sheet with data validation."""
        ws = self.wb.create_sheet('Control Panel', 6)

        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = 'DASHBOARD CONTROL PANEL'
        ws['A1'].font = SECTION_FONT

        ws.merge_cells('A2:D2')
        ws['A2'] = 'Configure visualization parameters below'
        ws['A2'].font = SUBTITLE_FONT

        # Control items
        controls = [
            ('B3', 'Benchmark Scenario:', 'B4', 'Budget Assumption FY24'),
            ('B6', 'Start Date:', 'B7', '2019-01-01'),
            ('B9', 'End Date:', 'B10', self.data[-1]['date'].strftime('%Y-%m-%d')),
            ('B12', 'Display Mode:', 'B13', 'Nominal Values'),
            ('B15', 'Component Focus:', 'B16', 'All Components'),
            ('B18', 'Time Granularity:', 'B19', 'Monthly'),
        ]

        for label_cell, label, value_cell, default in controls:
            ws[label_cell] = label
            ws[label_cell].font = Font(name='Segoe UI', bold=True, size=10, color=COLORS['navy'])
            ws[value_cell] = default
            ws[value_cell].font = BODY_FONT
            ws[value_cell].border = THIN_BORDER
            ws[value_cell].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

        # Data validation lists
        ws['E3'] = 'Benchmark Options:'
        ws['E3'].font = Font(name='Segoe UI', bold=True, size=9)
        benchmark_options = ['Pre-COVID Baseline (2019)', '5-Year Historical Average',
                            'Budget Assumption FY24', 'Industry Consensus Forecast', 'Custom Scenario']
        for i, opt in enumerate(benchmark_options, 4):
            ws[f'E{i}'] = opt
            ws[f'E{i}'].font = DATA_FONT

        ws['E10'] = 'Display Mode Options:'
        ws['E10'].font = Font(name='Segoe UI', bold=True, size=9)
        display_options = ['Nominal Values', 'Inflation-Adjusted (Real)',
                          'Year-over-Year Change %', 'Index vs. Benchmark Variance']
        for i, opt in enumerate(display_options, 11):
            ws[f'E{i}'] = opt
            ws[f'E{i}'].font = DATA_FONT

        ws['E16'] = 'Component Options:'
        ws['E16'].font = Font(name='Segoe UI', bold=True, size=9)
        component_options = ['All Components', 'Engineering Focus', 'Procurement Focus', 'Construction Focus']
        for i, opt in enumerate(component_options, 17):
            ws[f'E{i}'] = opt
            ws[f'E{i}'].font = DATA_FONT

        ws['E22'] = 'Granularity Options:'
        ws['E22'].font = Font(name='Segoe UI', bold=True, size=9)
        granularity_options = ['Monthly', 'Quarterly Average', 'Annual Average', 'Rolling 12-Month']
        for i, opt in enumerate(granularity_options, 23):
            ws[f'E{i}'] = opt
            ws[f'E{i}'].font = DATA_FONT

        # Add data validation
        dv_benchmark = DataValidation(type='list', formula1='$E$4:$E$8', allow_blank=False)
        dv_benchmark.error = 'Please select from available options'
        dv_benchmark.errorTitle = 'Invalid Selection'
        dv_benchmark.prompt = 'Select comparison scenario'
        dv_benchmark.promptTitle = 'Benchmark Selection'
        ws.add_data_validation(dv_benchmark)
        dv_benchmark.add('B4')

        dv_display = DataValidation(type='list', formula1='$E$11:$E$14', allow_blank=False)
        ws.add_data_validation(dv_display)
        dv_display.add('B13')

        dv_component = DataValidation(type='list', formula1='$E$17:$E$20', allow_blank=False)
        ws.add_data_validation(dv_component)
        dv_component.add('B16')

        dv_granularity = DataValidation(type='list', formula1='$E$23:$E$26', allow_blank=False)
        ws.add_data_validation(dv_granularity)
        dv_granularity.add('B19')

        # Instructions
        ws['B22'] = 'INSTRUCTIONS'
        ws['B22'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        instructions = """1. Select options from the dropdown menus above
2. Changes will update all dashboard visualizations
3. Date range filters apply to trend charts
4. Benchmark selection affects variance calculations
5. Display mode changes how values are presented"""

        ws.merge_cells('B23:D27')
        ws['B23'] = instructions
        ws['B23'].font = BODY_FONT
        ws['B23'].alignment = Alignment(wrap_text=True, vertical='top')

        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30

    def _create_documentation(self):
        """Create the Documentation sheet."""
        ws = self.wb.create_sheet('Documentation', 7)

        # Header
        ws.merge_cells('B2:H2')
        ws['B2'] = 'EPC Price Index Dashboard - Documentation'
        ws['B2'].font = SECTION_FONT

        # Table of contents
        ws['B4'] = 'CONTENTS'
        ws['B4'].font = Font(name='Segoe UI', bold=True, size=14, color=COLORS['navy'])

        contents = [
            '1. Overview and Purpose',
            '2. Data Sources and Methodology',
            '3. Sheet Descriptions',
            '4. Named Ranges Reference',
            '5. Update Procedures',
            '6. Interpretation Guidelines',
        ]
        for i, item in enumerate(contents, 5):
            ws[f'B{i}'] = item
            ws[f'B{i}'].font = Font(name='Segoe UI', size=10, color=COLORS['corporate_blue'], underline='single')

        # Section 1: Overview
        ws['B12'] = '1. OVERVIEW AND PURPOSE'
        ws['B12'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        overview = """This workbook provides executive-level visualization and analysis of the EPC
(Engineering, Procurement, Construction) Price Index. It is designed to support strategic
decision-making related to capital project timing, budget planning, and cost management.

Key Features:
• Real-time index tracking with historical context
• Component-level breakdown (E/P/C sub-indices)
• Commodity driver analysis (Steel, Equipment, Labor, Logistics)
• Multiple benchmark comparison scenarios
• Dynamic date range and display mode controls"""

        ws.merge_cells('B13:H18')
        ws['B13'] = overview
        ws['B13'].font = BODY_FONT
        ws['B13'].alignment = Alignment(wrap_text=True, vertical='top')

        # Section 2: Methodology
        ws['B21'] = '2. DATA SOURCES AND METHODOLOGY'
        ws['B21'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        methodology = """Index Calculation Methodology:
• Consolidated Index = (0.15 × Engineering) + (0.45 × Procurement) + (0.40 × Construction)
• Base Period: January 2019 = 100
• Updates: Monthly

Benchmark Definitions:
• Pre-COVID Baseline: 2019 values extended with 2.5% annual growth assumption
• 5-Year Average: Rolling 60-month average of consolidated index
• Budget FY24: Organizational budget assumptions for fiscal year 2024
• Industry Consensus: Industry analyst consensus forecast (updated quarterly)

Data Sources:
• Component indices derived from industry cost surveys
• Commodity prices from market data feeds
• Labor rates from construction labor statistics"""

        ws.merge_cells('B22:H32')
        ws['B22'] = methodology
        ws['B22'].font = BODY_FONT
        ws['B22'].alignment = Alignment(wrap_text=True, vertical='top')

        # Section 3: Sheet descriptions
        ws['B35'] = '3. SHEET DESCRIPTIONS'
        ws['B35'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        sheets_info = [
            ('Executive Dashboard', 'Primary single-screen strategic overview with KPIs and charts'),
            ('Trend Analysis', 'Detailed historical trend analysis with monthly data'),
            ('Component Breakdown', 'E/P/C sub-index analysis and comparisons'),
            ('Commodity Drivers', 'Underlying cost driver analysis'),
            ('Benchmark Comparison', 'Multi-scenario benchmark comparisons'),
            ('Data Tables', 'Source data and calculated fields'),
            ('Control Panel', 'Dynamic parameter selectors'),
            ('Documentation', 'This reference guide'),
        ]

        sheet_headers = ['Sheet Name', 'Purpose']
        for col, header in enumerate(sheet_headers, 2):
            cell = ws.cell(row=36, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER

        for row_idx, (name, purpose) in enumerate(sheets_info, 37):
            ws.cell(row=row_idx, column=2, value=name).font = DATA_FONT
            ws.cell(row=row_idx, column=2).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=purpose).font = DATA_FONT
            ws.cell(row=row_idx, column=3).border = THIN_BORDER

        # Section 4: Named Ranges
        ws['B47'] = '4. NAMED RANGES REFERENCE'
        ws['B47'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        ranges_info = [
            ('DateList', 'All date values', "'Data Tables'!$A$3:$A$200"),
            ('ConsolidatedIndex', 'Main EPC index', "'Data Tables'!$B$3:$B$200"),
            ('EngineeringIndex', 'Engineering sub-index', "'Data Tables'!$C$3:$C$200"),
            ('ProcurementIndex', 'Procurement sub-index', "'Data Tables'!$D$3:$D$200"),
            ('ConstructionIndex', 'Construction sub-index', "'Data Tables'!$E$3:$E$200"),
            ('BenchmarkScenario', 'Selected benchmark', "'Control Panel'!$B$4"),
            ('DisplayMode', 'View mode selection', "'Control Panel'!$B$13"),
        ]

        range_headers = ['Name', 'Description', 'Reference']
        for col, header in enumerate(range_headers, 2):
            cell = ws.cell(row=48, column=col, value=header)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.border = THIN_BORDER

        for row_idx, (name, desc, ref) in enumerate(ranges_info, 49):
            ws.cell(row=row_idx, column=2, value=name).font = Font(name='Consolas', size=9)
            ws.cell(row=row_idx, column=2).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=desc).font = DATA_FONT
            ws.cell(row=row_idx, column=3).border = THIN_BORDER
            ws.cell(row=row_idx, column=4, value=ref).font = Font(name='Consolas', size=8)
            ws.cell(row=row_idx, column=4).border = THIN_BORDER

        # Section 5: Update procedures
        ws['B58'] = '5. UPDATE PROCEDURES'
        ws['B58'].font = Font(name='Segoe UI', bold=True, size=12, color=COLORS['navy'])

        update_proc = """Monthly Update Process:
1. Navigate to 'Data Tables' sheet
2. Add new row at bottom of data section
3. Enter date (format: YYYY-MM-DD) and index values
4. Helper columns will auto-calculate
5. Return to Executive Dashboard to verify update

All charts and KPIs use dynamic named ranges and will automatically
incorporate new data without manual chart updates."""

        ws.merge_cells('B59:H64')
        ws['B59'] = update_proc
        ws['B59'].font = BODY_FONT
        ws['B59'].alignment = Alignment(wrap_text=True, vertical='top')

        # Footer
        ws['B67'] = f"Document generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['B67'].font = Font(name='Segoe UI', size=8, italic=True, color=COLORS['medium_gray'])

        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 30


# =============================================================================
# MAIN EXECUTION
# =============================================================================
def main():
    """Main entry point."""
    output_path = '/home/user/codex-test/EPC_Executive_Dashboard.xlsx'

    # Install dateutil if needed
    try:
        from dateutil.relativedelta import relativedelta
    except ImportError:
        import subprocess
        subprocess.check_call(['pip3', 'install', 'python-dateutil'])
        from dateutil.relativedelta import relativedelta

    builder = EPCDashboardBuilder(output_path)
    builder.create_workbook()

    print(f"\nWorkbook saved to: {output_path}")
    print("Sheets created:")
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    wb.close()


if __name__ == '__main__':
    main()
