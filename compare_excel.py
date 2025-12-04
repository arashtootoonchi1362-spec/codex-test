#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel File Comparison Script for Persian Documents
Compares employer.xlsx and contractor.xlsx with advanced matching and RTL support.

Author: Lead Python Data Engineer
"""

import sys
import re
from typing import Optional, Tuple, List, Dict, Any
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from thefuzz import fuzz, process
import xlsxwriter


# =============================================================================
# CONFIGURATION
# =============================================================================

class Config:
    """Configuration constants for the comparison script."""

    # File paths
    EMPLOYER_FILE = "employer.xlsx"
    CONTRACTOR_FILE = "contractor.xlsx"
    OUTPUT_FILE = "disputes.xlsx"

    # Target sheet names
    EMPLOYER_SHEET = "ارزیابی مشاور زیربنایی"
    CONTRACTOR_SHEET = "ارزیابی پیمان‌کار"

    # Fuzzy matching thresholds
    SHEET_MATCH_THRESHOLD = 70
    COLUMN_MATCH_THRESHOLD = 75

    # Numeric comparison tolerance
    NUMERIC_TOLERANCE = 0.01

    # Potential ID column names (Persian and English)
    ID_COLUMN_CANDIDATES = [
        "ردیف", "شماره ردیف", "آیتم", "شناسه", "کد", "Row Number",
        "Item ID", "ID", "شماره", "ش", "row", "id", "item"
    ]

    # Font settings for output
    OUTPUT_FONT = "Tahoma"  # Alternative: "B Nazanin"


# =============================================================================
# PERSIAN TEXT NORMALIZATION
# =============================================================================

class PersianNormalizer:
    """Handles Persian character normalization and text cleaning."""

    # Character mapping for normalization
    CHAR_MAP = {
        'ي': 'ی',  # Arabic Yeh -> Persian Yeh
        'ك': 'ک',  # Arabic Kaf -> Persian Kaf
        '٠': '0', '١': '1', '٢': '2', '٣': '3', '٤': '4',
        '٥': '5', '٦': '6', '٧': '7', '٨': '8', '٩': '9',
        '۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4',
        '۵': '5', '۶': '6', '۷': '7', '۸': '8', '۹': '9',
    }

    # Zero-Width characters to remove
    ZERO_WIDTH_CHARS = [
        '\u200c',  # ZWNJ (Zero-Width Non-Joiner)
        '\u200b',  # Zero-Width Space
        '\u200d',  # ZWJ (Zero-Width Joiner)
        '\ufeff',  # BOM
        '\u200e',  # Left-to-Right Mark
        '\u200f',  # Right-to-Left Mark
    ]

    @classmethod
    def normalize(cls, text: Any) -> str:
        """
        Normalize Persian text by unifying characters and removing invisible chars.

        Args:
            text: Input text (can be any type, will be converted to string)

        Returns:
            Normalized string
        """
        if pd.isna(text):
            return ""

        text = str(text)

        # Apply character mapping
        for old_char, new_char in cls.CHAR_MAP.items():
            text = text.replace(old_char, new_char)

        # Remove zero-width characters
        for zw_char in cls.ZERO_WIDTH_CHARS:
            text = text.replace(zw_char, '')

        # Strip whitespace and normalize multiple spaces
        text = re.sub(r'\s+', ' ', text).strip()

        return text

    @classmethod
    def normalize_for_comparison(cls, text: Any) -> str:
        """
        Normalize text specifically for comparison (more aggressive).
        Removes all spaces for fuzzy matching.
        """
        normalized = cls.normalize(text)
        return normalized.replace(' ', '').lower()


# =============================================================================
# SHEET FINDER WITH FUZZY MATCHING
# =============================================================================

class SheetFinder:
    """Handles finding sheets with fuzzy matching support."""

    @staticmethod
    def get_available_sheets(file_path: str) -> List[str]:
        """Get list of available sheet names from Excel file."""
        try:
            wb = load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            raise FileNotFoundError(f"Cannot read Excel file '{file_path}': {e}")

    @classmethod
    def find_sheet(cls, file_path: str, target_sheet: str) -> Tuple[str, bool]:
        """
        Find the target sheet, using fuzzy matching if exact match not found.

        Args:
            file_path: Path to Excel file
            target_sheet: Desired sheet name

        Returns:
            Tuple of (found_sheet_name, was_fuzzy_matched)
        """
        available_sheets = cls.get_available_sheets(file_path)

        # Normalize target sheet name
        normalized_target = PersianNormalizer.normalize(target_sheet)

        # Try exact match first (with normalization)
        for sheet in available_sheets:
            if PersianNormalizer.normalize(sheet) == normalized_target:
                return sheet, False

        # Try fuzzy matching
        print(f"\n⚠️  Sheet '{target_sheet}' not found in '{file_path}'")
        print(f"   Available sheets: {available_sheets}")

        # Normalize all sheet names for comparison
        normalized_sheets = {
            PersianNormalizer.normalize(s): s for s in available_sheets
        }

        # Find best fuzzy match
        best_match = process.extractOne(
            normalized_target,
            list(normalized_sheets.keys()),
            scorer=fuzz.ratio
        )

        if best_match and best_match[1] >= Config.SHEET_MATCH_THRESHOLD:
            matched_sheet = normalized_sheets[best_match[0]]
            print(f"   ✓ Fuzzy matched to: '{matched_sheet}' (score: {best_match[1]}%)")
            return matched_sheet, True

        # If no good match, use the first sheet
        fallback_sheet = available_sheets[0]
        print(f"   ⚠️  No good match found. Using first sheet: '{fallback_sheet}'")
        return fallback_sheet, True


# =============================================================================
# COLUMN MAPPER WITH FUZZY MATCHING
# =============================================================================

class ColumnMapper:
    """Maps columns between two DataFrames using fuzzy matching."""

    def __init__(self, employer_columns: List[str], contractor_columns: List[str]):
        self.employer_columns = list(employer_columns)
        self.contractor_columns = list(contractor_columns)
        self.mapping: Dict[str, str] = {}
        self._build_mapping()

    def _build_mapping(self):
        """Build column mapping using fuzzy matching."""
        # Normalize column names
        employer_normalized = {
            PersianNormalizer.normalize_for_comparison(col): col
            for col in self.employer_columns
        }
        contractor_normalized = {
            PersianNormalizer.normalize_for_comparison(col): col
            for col in self.contractor_columns
        }

        # First pass: exact matches after normalization
        used_contractor_cols = set()
        for emp_norm, emp_orig in employer_normalized.items():
            if emp_norm in contractor_normalized:
                self.mapping[emp_orig] = contractor_normalized[emp_norm]
                used_contractor_cols.add(contractor_normalized[emp_norm])

        # Second pass: fuzzy matching for unmatched columns
        unmatched_employer = [
            col for col in self.employer_columns
            if col not in self.mapping
        ]
        unmatched_contractor = [
            col for col in self.contractor_columns
            if col not in used_contractor_cols
        ]

        if unmatched_employer and unmatched_contractor:
            contractor_norm_map = {
                PersianNormalizer.normalize_for_comparison(col): col
                for col in unmatched_contractor
            }

            for emp_col in unmatched_employer:
                emp_norm = PersianNormalizer.normalize_for_comparison(emp_col)
                match = process.extractOne(
                    emp_norm,
                    list(contractor_norm_map.keys()),
                    scorer=fuzz.ratio
                )

                if match and match[1] >= Config.COLUMN_MATCH_THRESHOLD:
                    matched_contractor = contractor_norm_map[match[0]]
                    self.mapping[emp_col] = matched_contractor
                    del contractor_norm_map[match[0]]

    def get_contractor_column(self, employer_column: str) -> Optional[str]:
        """Get the mapped contractor column for an employer column."""
        return self.mapping.get(employer_column)

    def get_mapped_columns(self) -> List[Tuple[str, str]]:
        """Get list of (employer_col, contractor_col) tuples."""
        return list(self.mapping.items())

    def print_mapping(self):
        """Print the column mapping for debugging."""
        print("\n📊 Column Mapping:")
        print("   " + "-" * 60)
        for emp_col, con_col in self.mapping.items():
            print(f"   Employer: '{emp_col}' <-> Contractor: '{con_col}'")
        print("   " + "-" * 60)


# =============================================================================
# VALUE COMPARATOR
# =============================================================================

class ValueComparator:
    """Handles comparison of values with type-aware logic."""

    @staticmethod
    def is_numeric(value: Any) -> bool:
        """Check if a value can be treated as numeric."""
        if pd.isna(value):
            return False
        if isinstance(value, (int, float, np.integer, np.floating)):
            return True
        try:
            float(str(value).replace(',', '').replace('٬', ''))
            return True
        except (ValueError, TypeError):
            return False

    @staticmethod
    def to_numeric(value: Any) -> float:
        """Convert value to numeric, handling Persian number separators."""
        if pd.isna(value):
            return 0.0
        if isinstance(value, (int, float, np.integer, np.floating)):
            return float(value)
        try:
            # Remove thousand separators (both Western and Persian)
            cleaned = str(value).replace(',', '').replace('٬', '').replace(' ', '')
            return float(cleaned)
        except (ValueError, TypeError):
            return 0.0

    @classmethod
    def compare(cls, val1: Any, val2: Any) -> Tuple[bool, Optional[float]]:
        """
        Compare two values.

        Args:
            val1: First value (contractor)
            val2: Second value (employer)

        Returns:
            Tuple of (are_equal, numeric_difference_or_none)
        """
        # Handle both being NaN/None/empty
        is_empty1 = pd.isna(val1) or str(val1).strip() == ""
        is_empty2 = pd.isna(val2) or str(val2).strip() == ""

        if is_empty1 and is_empty2:
            return True, None

        if is_empty1 != is_empty2:
            # One is empty, one is not
            if cls.is_numeric(val1) or cls.is_numeric(val2):
                return False, cls.to_numeric(val1) - cls.to_numeric(val2)
            return False, None

        # Both have values - check if numeric
        if cls.is_numeric(val1) and cls.is_numeric(val2):
            num1 = cls.to_numeric(val1)
            num2 = cls.to_numeric(val2)
            difference = num1 - num2

            if abs(difference) <= Config.NUMERIC_TOLERANCE:
                return True, None
            return False, difference

        # Text comparison with normalization
        norm1 = PersianNormalizer.normalize(val1)
        norm2 = PersianNormalizer.normalize(val2)

        return norm1 == norm2, None


# =============================================================================
# ROW ALIGNER
# =============================================================================

class RowAligner:
    """Handles row alignment between two DataFrames."""

    @staticmethod
    def find_id_column(df: pd.DataFrame) -> Optional[str]:
        """Find the ID/Row Number column in a DataFrame."""
        columns_normalized = {
            PersianNormalizer.normalize_for_comparison(col): col
            for col in df.columns
        }

        for candidate in Config.ID_COLUMN_CANDIDATES:
            candidate_norm = PersianNormalizer.normalize_for_comparison(candidate)

            # Exact match
            if candidate_norm in columns_normalized:
                return columns_normalized[candidate_norm]

            # Fuzzy match
            for col_norm, col_orig in columns_normalized.items():
                if fuzz.ratio(candidate_norm, col_norm) >= 80:
                    return col_orig

        return None

    @classmethod
    def align_dataframes(
        cls,
        df_employer: pd.DataFrame,
        df_contractor: pd.DataFrame
    ) -> Tuple[pd.DataFrame, pd.DataFrame, Optional[str]]:
        """
        Align two DataFrames by ID column or index.

        Returns:
            Tuple of (aligned_employer, aligned_contractor, id_column_name)
        """
        # Find ID columns
        emp_id_col = cls.find_id_column(df_employer)
        con_id_col = cls.find_id_column(df_contractor)

        if emp_id_col and con_id_col:
            print(f"\n🔑 Using ID columns for alignment:")
            print(f"   Employer ID column: '{emp_id_col}'")
            print(f"   Contractor ID column: '{con_id_col}'")

            # Normalize ID values
            df_employer = df_employer.copy()
            df_contractor = df_contractor.copy()

            df_employer['_align_key'] = df_employer[emp_id_col].apply(
                PersianNormalizer.normalize
            )
            df_contractor['_align_key'] = df_contractor[con_id_col].apply(
                PersianNormalizer.normalize
            )

            # Set index for alignment
            df_employer = df_employer.set_index('_align_key')
            df_contractor = df_contractor.set_index('_align_key')

            # Get common indices
            common_idx = df_employer.index.intersection(df_contractor.index)

            df_employer = df_employer.loc[common_idx].reset_index(drop=True)
            df_contractor = df_contractor.loc[common_idx].reset_index(drop=True)

            return df_employer, df_contractor, emp_id_col

        print("\n📋 No ID column found. Aligning by row index.")

        # Align by index - use minimum length
        min_len = min(len(df_employer), len(df_contractor))
        return (
            df_employer.head(min_len).reset_index(drop=True),
            df_contractor.head(min_len).reset_index(drop=True),
            None
        )


# =============================================================================
# DISPUTE REPORT GENERATOR
# =============================================================================

class DisputeReport:
    """Generates the dispute report with RTL formatting."""

    def __init__(self, output_path: str):
        self.output_path = output_path
        self.disputes: List[Dict[str, Any]] = []

    def add_dispute(
        self,
        row_identifier: str,
        column_name: str,
        contractor_value: Any,
        employer_value: Any,
        difference: Optional[float]
    ):
        """Add a dispute record."""
        self.disputes.append({
            'شناسه/ردیف': row_identifier,
            'نام ستون': column_name,
            'مقدار پیمانکار': contractor_value if not pd.isna(contractor_value) else "(خالی)",
            'مقدار کارفرما': employer_value if not pd.isna(employer_value) else "(خالی)",
            'اختلاف': difference if difference is not None else "-"
        })

    def generate(self) -> bool:
        """
        Generate the dispute report Excel file.

        Returns:
            True if disputes were found and file was created, False otherwise.
        """
        if not self.disputes:
            print("\n✅ No disputes found! The files match perfectly.")
            return False

        print(f"\n⚠️  Found {len(self.disputes)} dispute(s)!")

        # Create DataFrame
        df = pd.DataFrame(self.disputes)

        # Write to Excel with xlsxwriter for RTL support
        with pd.ExcelWriter(
            self.output_path,
            engine='xlsxwriter',
            engine_kwargs={'options': {'strings_to_urls': False}}
        ) as writer:
            df.to_excel(writer, sheet_name='اختلافات', index=False)

            workbook = writer.book
            worksheet = writer.sheets['اختلافات']

            # Set RTL
            worksheet.right_to_left()

            # Define formats
            header_format = workbook.add_format({
                'font_name': Config.OUTPUT_FONT,
                'font_size': 12,
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'text_wrap': True
            })

            cell_format = workbook.add_format({
                'font_name': Config.OUTPUT_FONT,
                'font_size': 11,
                'align': 'center',
                'valign': 'vcenter',
                'border': 1
            })

            dispute_format = workbook.add_format({
                'font_name': Config.OUTPUT_FONT,
                'font_size': 11,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#FF6B6B',  # Red background
                'font_color': 'white',
                'border': 1,
                'bold': True
            })

            difference_format = workbook.add_format({
                'font_name': Config.OUTPUT_FONT,
                'font_size': 11,
                'align': 'center',
                'valign': 'vcenter',
                'bg_color': '#FFE66D',  # Yellow background
                'border': 1,
                'num_format': '#,##0.00'
            })

            # Set column widths
            column_widths = [20, 25, 25, 25, 18]
            for col_idx, width in enumerate(column_widths):
                worksheet.set_column(col_idx, col_idx, width)

            # Write headers with formatting
            for col_idx, col_name in enumerate(df.columns):
                worksheet.write(0, col_idx, col_name, header_format)

            # Write data with formatting
            for row_idx in range(len(df)):
                for col_idx, col_name in enumerate(df.columns):
                    value = df.iloc[row_idx, col_idx]

                    if col_name in ['مقدار پیمانکار', 'مقدار کارفرما']:
                        worksheet.write(row_idx + 1, col_idx, value, dispute_format)
                    elif col_name == 'اختلاف':
                        if value != "-":
                            worksheet.write(row_idx + 1, col_idx, value, difference_format)
                        else:
                            worksheet.write(row_idx + 1, col_idx, value, cell_format)
                    else:
                        worksheet.write(row_idx + 1, col_idx, value, cell_format)

            # Set row height
            for row_idx in range(len(df) + 1):
                worksheet.set_row(row_idx, 25)

            # Freeze header row
            worksheet.freeze_panes(1, 0)

        print(f"📄 Dispute report saved to: {self.output_path}")
        return True


# =============================================================================
# MAIN COMPARISON ENGINE
# =============================================================================

class ExcelComparator:
    """Main comparison engine orchestrating the full comparison process."""

    def __init__(
        self,
        employer_file: str = Config.EMPLOYER_FILE,
        contractor_file: str = Config.CONTRACTOR_FILE,
        output_file: str = Config.OUTPUT_FILE
    ):
        self.employer_file = employer_file
        self.contractor_file = contractor_file
        self.output_file = output_file

        self.df_employer: Optional[pd.DataFrame] = None
        self.df_contractor: Optional[pd.DataFrame] = None
        self.column_mapper: Optional[ColumnMapper] = None
        self.id_column: Optional[str] = None

    def _validate_files(self):
        """Validate that input files exist."""
        for file_path in [self.employer_file, self.contractor_file]:
            if not Path(file_path).exists():
                raise FileNotFoundError(f"File not found: {file_path}")
        print("✓ Input files validated")

    def _load_data(self):
        """Load data from both Excel files with smart sheet selection."""
        print("\n" + "=" * 70)
        print("📂 LOADING DATA")
        print("=" * 70)

        # Find and load employer sheet
        emp_sheet, emp_fuzzy = SheetFinder.find_sheet(
            self.employer_file, Config.EMPLOYER_SHEET
        )
        self.df_employer = pd.read_excel(
            self.employer_file,
            sheet_name=emp_sheet,
            engine='openpyxl'
        )
        print(f"   Loaded employer data: {len(self.df_employer)} rows, "
              f"{len(self.df_employer.columns)} columns")

        # Find and load contractor sheet
        con_sheet, con_fuzzy = SheetFinder.find_sheet(
            self.contractor_file, Config.CONTRACTOR_SHEET
        )
        self.df_contractor = pd.read_excel(
            self.contractor_file,
            sheet_name=con_sheet,
            engine='openpyxl'
        )
        print(f"   Loaded contractor data: {len(self.df_contractor)} rows, "
              f"{len(self.df_contractor.columns)} columns")

    def _prepare_data(self):
        """Prepare and align data for comparison."""
        print("\n" + "=" * 70)
        print("🔧 PREPARING DATA")
        print("=" * 70)

        # Build column mapping
        self.column_mapper = ColumnMapper(
            self.df_employer.columns,
            self.df_contractor.columns
        )
        self.column_mapper.print_mapping()

        # Align rows
        self.df_employer, self.df_contractor, self.id_column = RowAligner.align_dataframes(
            self.df_employer, self.df_contractor
        )

        print(f"\n   Aligned data: {len(self.df_employer)} rows to compare")

    def _compare_data(self) -> DisputeReport:
        """Perform the actual comparison and collect disputes."""
        print("\n" + "=" * 70)
        print("🔍 COMPARING DATA")
        print("=" * 70)

        report = DisputeReport(self.output_file)
        comparator = ValueComparator()

        mapped_columns = self.column_mapper.get_mapped_columns()
        total_comparisons = 0
        differences_found = 0

        for row_idx in range(len(self.df_employer)):
            # Get row identifier
            if self.id_column:
                row_id = str(self.df_employer.iloc[row_idx].get(self.id_column, row_idx + 1))
            else:
                row_id = str(row_idx + 1)

            for emp_col, con_col in mapped_columns:
                emp_value = self.df_employer.iloc[row_idx][emp_col]
                con_value = self.df_contractor.iloc[row_idx][con_col]

                total_comparisons += 1
                are_equal, difference = comparator.compare(con_value, emp_value)

                if not are_equal:
                    differences_found += 1
                    report.add_dispute(
                        row_identifier=row_id,
                        column_name=emp_col,
                        contractor_value=con_value,
                        employer_value=emp_value,
                        difference=difference
                    )

        print(f"\n   Total comparisons: {total_comparisons:,}")
        print(f"   Differences found: {differences_found:,}")

        return report

    def run(self) -> bool:
        """
        Run the full comparison process.

        Returns:
            True if process completed successfully, False otherwise.
        """
        print("\n" + "=" * 70)
        print("🚀 EXCEL COMPARISON TOOL - STARTING")
        print("=" * 70)
        print(f"   Employer file:   {self.employer_file}")
        print(f"   Contractor file: {self.contractor_file}")
        print(f"   Output file:     {self.output_file}")

        try:
            # Step 1: Validate files
            self._validate_files()

            # Step 2: Load data with smart sheet selection
            self._load_data()

            # Step 3: Prepare and align data
            self._prepare_data()

            # Step 4: Compare data
            report = self._compare_data()

            # Step 5: Generate report
            print("\n" + "=" * 70)
            print("📊 GENERATING REPORT")
            print("=" * 70)

            has_disputes = report.generate()

            print("\n" + "=" * 70)
            print("✅ COMPARISON COMPLETE")
            print("=" * 70)

            return True

        except FileNotFoundError as e:
            print(f"\n❌ Error: {e}")
            return False
        except Exception as e:
            print(f"\n❌ Unexpected error: {e}")
            import traceback
            traceback.print_exc()
            return False


# =============================================================================
# COMMAND LINE INTERFACE
# =============================================================================

def main():
    """Main entry point for the script."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Compare Excel files and generate dispute report",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python compare_excel.py
  python compare_excel.py --employer emp.xlsx --contractor con.xlsx
  python compare_excel.py -o report.xlsx
        """
    )

    parser.add_argument(
        '--employer', '-e',
        default=Config.EMPLOYER_FILE,
        help=f"Employer Excel file (default: {Config.EMPLOYER_FILE})"
    )

    parser.add_argument(
        '--contractor', '-c',
        default=Config.CONTRACTOR_FILE,
        help=f"Contractor Excel file (default: {Config.CONTRACTOR_FILE})"
    )

    parser.add_argument(
        '--output', '-o',
        default=Config.OUTPUT_FILE,
        help=f"Output disputes file (default: {Config.OUTPUT_FILE})"
    )

    parser.add_argument(
        '--tolerance', '-t',
        type=float,
        default=Config.NUMERIC_TOLERANCE,
        help=f"Numeric comparison tolerance (default: {Config.NUMERIC_TOLERANCE})"
    )

    args = parser.parse_args()

    # Update config if tolerance specified
    if args.tolerance != Config.NUMERIC_TOLERANCE:
        Config.NUMERIC_TOLERANCE = args.tolerance

    # Run comparison
    comparator = ExcelComparator(
        employer_file=args.employer,
        contractor_file=args.contractor,
        output_file=args.output
    )

    success = comparator.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
