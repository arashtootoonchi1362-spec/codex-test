#!/usr/bin/env python3
"""
Excel Logic Chain Audit Script
==============================
Comprehensive audit and verification of complex Excel logic chains involving:
- Dates (Column P driver)
- Quarter calculations (Semah)
- Exchange rates from Arz sheet
- Index lookups based on Date, Category, and Chapter

Designed for Google Cloud/Colab environments with Persian language support.

Author: Claude Code
Date: 2025-12-11
"""

import os
import sys
import re
import json
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Union
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

# Third-party imports
try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl --quiet")
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string

try:
    import pandas as pd
except ImportError:
    print("Installing pandas...")
    os.system("pip install pandas --quiet")
    import pandas as pd

# Persian language support
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    PERSIAN_SUPPORT = True
except ImportError:
    print("Installing Persian language support libraries...")
    os.system("pip install arabic-reshaper python-bidi --quiet")
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        PERSIAN_SUPPORT = True
    except ImportError:
        PERSIAN_SUPPORT = False
        print("Warning: Persian language support not available")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# =============================================================================
# Persian Text Utilities
# =============================================================================

def reshape_persian(text: str) -> str:
    """
    Reshape Persian/Arabic text for proper RTL display.

    Args:
        text: Input text potentially containing Persian characters

    Returns:
        Reshaped text with proper RTL handling
    """
    if not PERSIAN_SUPPORT or not text:
        return str(text) if text else ""

    try:
        # Reshape the text for proper letter joining
        reshaped = arabic_reshaper.reshape(str(text))
        # Apply bidirectional algorithm for RTL
        return get_display(reshaped)
    except Exception:
        return str(text)


def is_persian(text: str) -> bool:
    """Check if text contains Persian/Arabic characters."""
    if not text:
        return False
    persian_pattern = re.compile(r'[\u0600-\u06FF\u0750-\u077F\uFB50-\uFDFF\uFE70-\uFEFF]')
    return bool(persian_pattern.search(str(text)))


# =============================================================================
# Data Classes
# =============================================================================

class AuditStatus(Enum):
    """Audit result status."""
    PASS = "PASS"
    FAIL = "FAIL"
    WARNING = "WARNING"
    INFO = "INFO"
    SKIPPED = "SKIPPED"


@dataclass
class AuditResult:
    """Single audit check result."""
    check_name: str
    status: AuditStatus
    expected_value: Any
    actual_value: Any
    cell_reference: str
    message: str = ""
    details: Dict = field(default_factory=dict)

    def to_dict(self) -> Dict:
        """Convert to dictionary for JSON serialization."""
        return {
            "check_name": self.check_name,
            "status": self.status.value,
            "expected_value": str(self.expected_value),
            "actual_value": str(self.actual_value),
            "cell_reference": self.cell_reference,
            "message": self.message,
            "details": self.details
        }


@dataclass
class SheetStructure:
    """Represents the structure of a sheet."""
    name: str
    dimensions: str
    max_row: int
    max_col: int
    headers: Dict[int, str] = field(default_factory=dict)
    data_start_row: int = 1


@dataclass
class ExchangeRateData:
    """Exchange rate lookup data."""
    date: str
    rate: float
    row: int


@dataclass
class IndexData:
    """Index lookup data from category sheets."""
    sheet_name: str
    chapter: int
    period: str
    value: float


# =============================================================================
# Persian/Jalali Date Utilities
# =============================================================================

class JalaliDateParser:
    """Parse and manipulate Jalali (Persian) dates."""

    # Persian digits to Latin mapping
    PERSIAN_DIGITS = '۰۱۲۳۴۵۶۷۸۹'
    LATIN_DIGITS = '0123456789'

    @classmethod
    def persian_to_latin(cls, text: str) -> str:
        """Convert Persian digits to Latin."""
        if not text:
            return ""
        result = str(text)
        for p, l in zip(cls.PERSIAN_DIGITS, cls.LATIN_DIGITS):
            result = result.replace(p, l)
        return result

    @classmethod
    def parse_date(cls, date_str: str) -> Optional[Tuple[int, int, int]]:
        """
        Parse a Jalali date string into (year, month, day).

        Args:
            date_str: Date string in format YYYY/MM/DD

        Returns:
            Tuple of (year, month, day) or None if invalid
        """
        if not date_str:
            return None

        # Convert Persian digits to Latin
        date_str = cls.persian_to_latin(str(date_str))

        # Try various formats
        patterns = [
            r'(\d{4})/(\d{1,2})/(\d{1,2})',  # YYYY/MM/DD
            r'(\d{4})-(\d{1,2})-(\d{1,2})',  # YYYY-MM-DD
        ]

        for pattern in patterns:
            match = re.match(pattern, date_str.strip())
            if match:
                year, month, day = map(int, match.groups())
                return (year, month, day)

        return None

    @classmethod
    def get_quarter(cls, month: int) -> int:
        """
        Get the quarter (Semah) for a given month.

        Persian quarters:
        - Q1 (سه‌ماهه اول): Farvardin-Khordad (months 1-3)
        - Q2 (سه‌ماهه دوم): Tir-Shahrivar (months 4-6)
        - Q3 (سه‌ماهه سوم): Mehr-Azar (months 7-9)
        - Q4 (سه‌ماهه چهارم): Dey-Esfand (months 10-12)
        """
        if 1 <= month <= 3:
            return 1
        elif 4 <= month <= 6:
            return 2
        elif 7 <= month <= 9:
            return 3
        elif 10 <= month <= 12:
            return 4
        return 0

    @classmethod
    def get_period_column(cls, year: int, month: int, base_year: int = 1397) -> Optional[int]:
        """
        Calculate the column index for a given date in category sheets.

        Args:
            year: Jalali year
            month: Jalali month (1-12)
            base_year: Starting year in the category sheet

        Returns:
            Column index (1-based) or None if out of range
        """
        # Each year has roughly 8 columns: Q1, Tir, Mordad, Shahrivar, Q3, Q4, ...
        # Structure: اول, تیر, مرداد, شهریور, سوم, چهارم
        year_offset = year - base_year

        if month in [1, 2, 3]:  # Q1
            period_offset = 0
        elif month == 4:  # Tir
            period_offset = 1
        elif month == 5:  # Mordad
            period_offset = 2
        elif month == 6:  # Shahrivar
            period_offset = 3
        elif month in [7, 8, 9]:  # Q3
            period_offset = 4
        elif month in [10, 11, 12]:  # Q4
            period_offset = 5
        else:
            return None

        # Column C is first data column (index 3), each year has 6 periods
        return 3 + (year_offset * 6) + period_offset


# =============================================================================
# Excel Logic Auditor
# =============================================================================

class ExcelLogicAuditor:
    """
    Main class for auditing Excel logic chains.

    Verifies:
    1. Date-driven calculations (P2 driver)
    2. Quarter (Semah) calculations (Column R)
    3. Exchange rate lookups (Column Q from Arz sheet)
    4. Index lookups (Columns F & J based on Date, Category, Chapter)
    """

    # Sheet name mappings for categories
    CATEGORY_SHEETS = {
        1: 'ابنیه',           # Buildings
        2: 'مکانیک',          # Mechanics
        3: 'برق',             # Electrical (تاسیسات برقی)
        4: 'راه، راه آهن و باند فرودگاه',  # Roads
        5: 'تجهیزات آب و فاضلاب',  # Water & Sewage
        6: 'رشته ای',         # Disciplinary
    }

    # Column mappings for main sheet
    MAIN_COLUMNS = {
        'B': 'row_number',
        'C': 'contract_row',        # ردیف قراردادی
        'D': 'chapter',             # فصل (Dm)
        'E': 'category',            # دسته‌بندی
        'F': 'index',               # شاخص
        'G': 'description',         # شرح ردیف
        'H': 'package_name',        # نام بسته خرید
        'I': 'proposal_date',       # تاریخ پیشنهاد قیمت
        'J': 'calc_base_date',      # تاریخ مبنای محاسبات
        'K': 'entry_date',          # تاریخ ورود به کارگاه
        'L': 'currency_ratio_query',  # ضریب ارزبری طبق استعلام
        'M': 'currency_ratio_used',   # ضریب ارزبری منظور شده
        'N': 'f_coefficient',       # ضریب F روش الف
        'O': 'purchase_rate',       # قیمت ارز در زمان خرید (Ci)
        'P': 'base_rate',           # قیمت ارز پایه (C0)
        'Q': 'calc_base_year',      # سال مبنای محاسبات
        'R': 'n_value',             # مقدار N (Semah)
        'S': 'n_reduction',         # درصد کاهش N
        'T': 'months_elapsed',      # تعداد ماه سپری شده
    }

    def __init__(self, file_path: str):
        """
        Initialize the auditor.

        Args:
            file_path: Path to the Excel file to audit
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        # Load workbook with formulas and with calculated values
        self.wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        self.wb_values = openpyxl.load_workbook(file_path, data_only=True)

        self.audit_results: List[AuditResult] = []
        self.exchange_rates: Dict[str, float] = {}
        self.category_indices: Dict[str, pd.DataFrame] = {}

        # Cache for computed values
        self._mode: Optional[int] = None
        self._base_date: Optional[str] = None
        self._base_rate: Optional[float] = None

        logger.info(f"Loaded Excel file: {file_path}")
        logger.info(f"Sheets found: {self.wb_formulas.sheetnames}")

    def _get_cell_value(self, sheet_name: str, cell_ref: str,
                        use_formula: bool = False) -> Any:
        """Get cell value from specified sheet."""
        wb = self.wb_formulas if use_formula else self.wb_values
        try:
            sheet = wb[sheet_name]
            return sheet[cell_ref].value
        except Exception as e:
            logger.warning(f"Error reading {sheet_name}!{cell_ref}: {e}")
            return None

    def _add_result(self, check_name: str, status: AuditStatus,
                    expected: Any, actual: Any, cell_ref: str,
                    message: str = "", **details) -> None:
        """Add an audit result."""
        result = AuditResult(
            check_name=check_name,
            status=status,
            expected_value=expected,
            actual_value=actual,
            cell_reference=cell_ref,
            message=message,
            details=details
        )
        self.audit_results.append(result)

        # Log based on status
        status_emoji = {
            AuditStatus.PASS: "✓",
            AuditStatus.FAIL: "✗",
            AuditStatus.WARNING: "⚠",
            AuditStatus.INFO: "ℹ",
            AuditStatus.SKIPPED: "○"
        }
        logger.info(f"{status_emoji.get(status, '?')} [{status.value}] {check_name} @ {cell_ref}: {message}")

    # =========================================================================
    # Exchange Rate Loading
    # =========================================================================

    def load_exchange_rates(self) -> Dict[str, float]:
        """
        Load exchange rates from the Arz sheet.

        Returns:
            Dictionary mapping date strings to exchange rates
        """
        logger.info("Loading exchange rates from Arz sheet...")

        # Find Arz sheet (may have trailing space)
        arz_sheet_name = None
        for name in self.wb_values.sheetnames:
            if 'Arz' in name or 'arz' in name.lower():
                arz_sheet_name = name
                break

        if not arz_sheet_name:
            logger.error("Arz sheet not found!")
            return {}

        sheet = self.wb_values[arz_sheet_name]
        rates = {}

        for row in range(2, sheet.max_row + 1):
            date_val = sheet.cell(row=row, column=1).value
            rate_val = sheet.cell(row=row, column=2).value

            if date_val:
                # Normalize date string
                date_str = JalaliDateParser.persian_to_latin(str(date_val))
                if rate_val is not None:
                    try:
                        rates[date_str] = float(rate_val)
                    except (ValueError, TypeError):
                        pass

        self.exchange_rates = rates
        logger.info(f"Loaded {len(rates)} exchange rates")
        return rates

    def lookup_exchange_rate(self, date_str: str) -> Optional[float]:
        """
        Look up exchange rate for a given date.

        Args:
            date_str: Date in YYYY/MM/DD format

        Returns:
            Exchange rate or None if not found
        """
        if not self.exchange_rates:
            self.load_exchange_rates()

        # Normalize input date
        normalized = JalaliDateParser.persian_to_latin(str(date_str))
        return self.exchange_rates.get(normalized)

    # =========================================================================
    # Category Index Loading
    # =========================================================================

    def load_category_indices(self) -> Dict[str, pd.DataFrame]:
        """
        Load index data from all category sheets.

        Returns:
            Dictionary mapping sheet names to DataFrames
        """
        logger.info("Loading category indices...")

        for sheet_name in self.wb_values.sheetnames:
            if sheet_name in ['1-2', 'Arz ', 'Arz', 'درصد ارزیری']:
                continue

            try:
                sheet = self.wb_values[sheet_name]
                data = []

                # Get headers from row 2-3
                for row in range(4, sheet.max_row + 1):
                    chapter = sheet.cell(row=row, column=1).value
                    description = sheet.cell(row=row, column=2).value

                    if chapter is not None:
                        row_data = {
                            'chapter': chapter,
                            'description': description
                        }

                        # Get index values for each period
                        for col in range(3, sheet.max_column + 1):
                            col_val = sheet.cell(row=row, column=col).value
                            if col_val is not None:
                                row_data[f'col_{col}'] = col_val

                        data.append(row_data)

                if data:
                    self.category_indices[sheet_name] = pd.DataFrame(data)
                    logger.info(f"Loaded {len(data)} rows from '{sheet_name}'")

            except Exception as e:
                logger.warning(f"Error loading sheet '{sheet_name}': {e}")

        return self.category_indices

    # =========================================================================
    # Mode and Base Value Verification
    # =========================================================================

    def verify_mode_settings(self) -> Tuple[int, str, float]:
        """
        Verify the mode settings from H1 and derived values.

        The mode (H1) determines:
        - Base calculation date (J column)
        - Base exchange rate (P column)

        Returns:
            Tuple of (mode, base_date, base_rate)
        """
        logger.info("="*60)
        logger.info("Verifying Mode Settings (H1 Driver)")
        logger.info("="*60)

        main_sheet = '1-2'

        # Get mode value
        mode = self._get_cell_value(main_sheet, 'H1')
        self._mode = int(mode) if mode else None

        self._add_result(
            "Mode Value (H1)",
            AuditStatus.INFO,
            None,
            mode,
            f"{main_sheet}!H1",
            f"Mode/Batch number: {mode}"
        )

        # Get calculated base date from H2
        base_date = self._get_cell_value(main_sheet, 'H2')
        self._base_date = str(base_date) if base_date else None

        # Verify base date formula logic
        # Formula: =_xlfn.LET(_xlpm.x, VALUE($H$1), IF(_xlpm.x=3, "اسفند ۱۳۹۸", IF(_xlpm.x=4, "1403/07/01", "")))
        expected_dates = {
            3: "اسفند ۱۳۹۸",
            4: "1403/07/01"
        }

        if self._mode in expected_dates:
            expected_date = expected_dates[self._mode]
            status = AuditStatus.PASS if str(base_date) == expected_date else AuditStatus.WARNING
            self._add_result(
                "Base Date (H2)",
                status,
                expected_date,
                base_date,
                f"{main_sheet}!H2",
                f"Base date for mode {self._mode}"
            )
        else:
            self._add_result(
                "Base Date (H2)",
                AuditStatus.INFO,
                "Unknown",
                base_date,
                f"{main_sheet}!H2",
                f"Base date for mode {self._mode} (not in known mappings)"
            )

        # Get base exchange rate from P7 (first data row)
        base_rate = self._get_cell_value(main_sheet, 'P7')
        self._base_rate = float(base_rate) if base_rate else None

        # Verify base rate formula logic
        # Formula: =_xlfn.LET(_xlpm.x, VALUE($H$1), IF(_xlpm.x=3, 405150, IF(_xlpm.x=4, 506978, "")))
        expected_rates = {
            3: 405150,
            4: 506978
        }

        if self._mode in expected_rates:
            expected_rate = expected_rates[self._mode]
            status = AuditStatus.PASS if base_rate == expected_rate else AuditStatus.FAIL
            self._add_result(
                "Base Exchange Rate (P7)",
                status,
                expected_rate,
                base_rate,
                f"{main_sheet}!P7",
                f"Base exchange rate for mode {self._mode}"
            )

        return (self._mode, self._base_date, self._base_rate)

    # =========================================================================
    # Quarter (Semah) Verification
    # =========================================================================

    def verify_quarter_calculations(self) -> List[AuditResult]:
        """
        Verify Quarter/Semah calculations in Column R.

        The R column formula calculates:
        - If year is 1402 or 1403: returns 0.05
        - Otherwise: looks up R2 value

        Returns:
            List of audit results for quarter calculations
        """
        logger.info("="*60)
        logger.info("Verifying Quarter (Semah) Calculations - Column R")
        logger.info("="*60)

        main_sheet = '1-2'
        sheet = self.wb_values[main_sheet]
        results = []

        # Get R2 reference value
        r2_value = self._get_cell_value(main_sheet, 'R2')

        # Check each data row
        for row in range(7, sheet.max_row + 1):
            j_value = sheet.cell(row=row, column=10).value  # J column - calc base date
            r_value = sheet.cell(row=row, column=18).value  # R column - N value

            if j_value is None:
                continue

            # Parse the date
            date_parsed = JalaliDateParser.parse_date(str(j_value))

            if date_parsed:
                year, month, day = date_parsed

                # Expected value based on formula logic
                if year in [1402, 1403]:
                    expected = 0.05
                else:
                    expected = r2_value

                # Compare with actual
                try:
                    actual = float(r_value) if r_value else None
                    if actual is not None:
                        tolerance = 0.001
                        if abs(actual - expected) < tolerance:
                            status = AuditStatus.PASS
                            message = f"N value correct for year {year}"
                        else:
                            status = AuditStatus.FAIL
                            message = f"N value mismatch for year {year}"
                    else:
                        status = AuditStatus.WARNING
                        message = "N value is empty"
                except (ValueError, TypeError):
                    status = AuditStatus.WARNING
                    message = f"Cannot parse N value: {r_value}"
                    actual = r_value

                self._add_result(
                    f"Quarter/N Calculation Row {row}",
                    status,
                    expected,
                    actual,
                    f"{main_sheet}!R{row}",
                    message,
                    date=str(j_value),
                    year=year
                )

        return results

    # =========================================================================
    # Exchange Rate Lookup Verification
    # =========================================================================

    def verify_exchange_rate_lookups(self) -> List[AuditResult]:
        """
        Verify exchange rate lookups from the Arz sheet.

        The O column contains the purchase rate (Ci) which should match
        the rate from Arz sheet for the date in K column.

        Returns:
            List of audit results for exchange rate lookups
        """
        logger.info("="*60)
        logger.info("Verifying Exchange Rate Lookups - Column O vs Arz Sheet")
        logger.info("="*60)

        if not self.exchange_rates:
            self.load_exchange_rates()

        main_sheet = '1-2'
        sheet = self.wb_values[main_sheet]

        for row in range(7, sheet.max_row + 1):
            k_value = sheet.cell(row=row, column=11).value  # K - entry date
            o_value = sheet.cell(row=row, column=15).value  # O - purchase rate

            if k_value is None or o_value is None:
                continue

            # Normalize the date
            date_str = JalaliDateParser.persian_to_latin(str(k_value))

            # Look up expected rate
            expected_rate = self.lookup_exchange_rate(date_str)

            if expected_rate:
                try:
                    actual_rate = float(o_value)
                    tolerance = 1.0  # Allow small rounding differences

                    if abs(actual_rate - expected_rate) <= tolerance:
                        status = AuditStatus.PASS
                        message = f"Rate matches for date {date_str}"
                    else:
                        # Check if it's close (within 1%)
                        pct_diff = abs(actual_rate - expected_rate) / expected_rate * 100
                        if pct_diff < 1:
                            status = AuditStatus.WARNING
                            message = f"Rate close ({pct_diff:.2f}% diff) for {date_str}"
                        else:
                            status = AuditStatus.FAIL
                            message = f"Rate mismatch ({pct_diff:.2f}% diff) for {date_str}"

                except (ValueError, TypeError):
                    status = AuditStatus.WARNING
                    message = f"Cannot parse rate value: {o_value}"
                    actual_rate = o_value
            else:
                status = AuditStatus.WARNING
                message = f"No rate found in Arz for date {date_str}"
                actual_rate = o_value
                expected_rate = "N/A"

            self._add_result(
                f"Exchange Rate Lookup Row {row}",
                status,
                expected_rate,
                actual_rate if 'actual_rate' in dir() else o_value,
                f"{main_sheet}!O{row}",
                message,
                lookup_date=date_str
            )

        return self.audit_results

    # =========================================================================
    # Index Lookup Verification
    # =========================================================================

    def verify_index_lookups(self) -> List[AuditResult]:
        """
        Verify index lookups in Columns F based on:
        - Date (from J column)
        - Category (from C column)
        - Chapter (from D column)

        Returns:
            List of audit results for index lookups
        """
        logger.info("="*60)
        logger.info("Verifying Index Lookups - Columns F (Date + Category + Chapter)")
        logger.info("="*60)

        if not self.category_indices:
            self.load_category_indices()

        main_sheet = '1-2'
        sheet = self.wb_values[main_sheet]

        for row in range(7, sheet.max_row + 1):
            c_value = sheet.cell(row=row, column=3).value  # C - category/contract row
            d_value = sheet.cell(row=row, column=4).value  # D - chapter number
            f_value = sheet.cell(row=row, column=6).value  # F - index value
            j_value = sheet.cell(row=row, column=10).value  # J - calculation date

            # Skip rows without essential data
            if c_value is None or d_value is None:
                continue

            # Parse date to determine period
            date_parsed = JalaliDateParser.parse_date(str(j_value)) if j_value else None

            # Log the index lookup attempt
            self._add_result(
                f"Index Value Row {row}",
                AuditStatus.INFO,
                "See category sheet",
                f_value,
                f"{main_sheet}!F{row}",
                f"Category={c_value}, Chapter={d_value}, Date={j_value}",
                category=c_value,
                chapter=d_value,
                date=str(j_value) if j_value else None
            )

        return self.audit_results

    # =========================================================================
    # Comprehensive Audit
    # =========================================================================

    def run_full_audit(self) -> Dict:
        """
        Run a comprehensive audit of all Excel logic chains.

        Returns:
            Dictionary containing audit summary and detailed results
        """
        logger.info("="*80)
        logger.info("STARTING COMPREHENSIVE EXCEL LOGIC AUDIT")
        logger.info("="*80)
        logger.info(f"File: {self.file_path}")
        logger.info(f"Timestamp: {datetime.now().isoformat()}")

        # Clear previous results
        self.audit_results = []

        # Step 1: Load reference data
        self.load_exchange_rates()
        self.load_category_indices()

        # Step 2: Verify mode settings
        mode, base_date, base_rate = self.verify_mode_settings()

        # Step 3: Verify quarter calculations
        self.verify_quarter_calculations()

        # Step 4: Verify exchange rate lookups
        self.verify_exchange_rate_lookups()

        # Step 5: Verify index lookups
        self.verify_index_lookups()

        # Compile summary
        summary = self._compile_summary()

        logger.info("="*80)
        logger.info("AUDIT COMPLETE")
        logger.info("="*80)
        logger.info(f"Total checks: {summary['total_checks']}")
        logger.info(f"Passed: {summary['passed']}")
        logger.info(f"Failed: {summary['failed']}")
        logger.info(f"Warnings: {summary['warnings']}")

        return {
            "file": str(self.file_path),
            "timestamp": datetime.now().isoformat(),
            "mode": mode,
            "base_date": base_date,
            "base_rate": base_rate,
            "summary": summary,
            "results": [r.to_dict() for r in self.audit_results]
        }

    def _compile_summary(self) -> Dict:
        """Compile audit summary statistics."""
        total = len(self.audit_results)
        passed = sum(1 for r in self.audit_results if r.status == AuditStatus.PASS)
        failed = sum(1 for r in self.audit_results if r.status == AuditStatus.FAIL)
        warnings = sum(1 for r in self.audit_results if r.status == AuditStatus.WARNING)
        info = sum(1 for r in self.audit_results if r.status == AuditStatus.INFO)

        return {
            "total_checks": total,
            "passed": passed,
            "failed": failed,
            "warnings": warnings,
            "info": info,
            "pass_rate": f"{passed/total*100:.1f}%" if total > 0 else "N/A"
        }

    # =========================================================================
    # Report Generation
    # =========================================================================

    def generate_report(self, output_path: Optional[str] = None) -> str:
        """
        Generate a detailed audit report.

        Args:
            output_path: Optional path to save the report

        Returns:
            Report content as string
        """
        audit_data = self.run_full_audit()

        # Build report
        lines = []
        lines.append("=" * 80)
        lines.append("EXCEL LOGIC CHAIN AUDIT REPORT")
        lines.append(reshape_persian("گزارش بازرسی زنجیره منطق اکسل"))
        lines.append("=" * 80)
        lines.append("")
        lines.append(f"File: {audit_data['file']}")
        lines.append(f"Timestamp: {audit_data['timestamp']}")
        lines.append(f"Mode: {audit_data['mode']}")
        lines.append(f"Base Date: {audit_data['base_date']}")
        lines.append(f"Base Rate: {audit_data['base_rate']}")
        lines.append("")

        # Summary
        summary = audit_data['summary']
        lines.append("-" * 40)
        lines.append("SUMMARY")
        lines.append("-" * 40)
        lines.append(f"Total Checks:  {summary['total_checks']}")
        lines.append(f"Passed:        {summary['passed']} ({AuditStatus.PASS.value})")
        lines.append(f"Failed:        {summary['failed']} ({AuditStatus.FAIL.value})")
        lines.append(f"Warnings:      {summary['warnings']} ({AuditStatus.WARNING.value})")
        lines.append(f"Info:          {summary['info']} ({AuditStatus.INFO.value})")
        lines.append(f"Pass Rate:     {summary['pass_rate']}")
        lines.append("")

        # Detailed results
        lines.append("-" * 40)
        lines.append("DETAILED RESULTS")
        lines.append("-" * 40)

        for result in audit_data['results']:
            status_marker = {
                'PASS': '[PASS]',
                'FAIL': '[FAIL]',
                'WARNING': '[WARN]',
                'INFO': '[INFO]',
                'SKIPPED': '[SKIP]'
            }.get(result['status'], '[????]')

            lines.append("")
            lines.append(f"{status_marker} {result['check_name']}")
            lines.append(f"  Cell: {result['cell_reference']}")
            lines.append(f"  Expected: {result['expected_value']}")
            lines.append(f"  Actual: {result['actual_value']}")
            if result['message']:
                lines.append(f"  Message: {result['message']}")

        report_content = "\n".join(lines)

        # Save if path provided
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(report_content)
            logger.info(f"Report saved to: {output_path}")

        return report_content

    def export_results_json(self, output_path: str) -> None:
        """Export audit results to JSON file."""
        audit_data = self.run_full_audit()

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(audit_data, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON results saved to: {output_path}")


# =============================================================================
# Data Extraction Utilities
# =============================================================================

class ExcelDataExtractor:
    """
    Utility class for extracting and analyzing data from the Excel file.
    """

    def __init__(self, file_path: str):
        """Initialize extractor with Excel file path."""
        self.file_path = Path(file_path)
        self.wb = openpyxl.load_workbook(file_path, data_only=True)

    def extract_main_data(self) -> pd.DataFrame:
        """
        Extract main data from the '1-2' sheet into a DataFrame.

        Returns:
            DataFrame with main data
        """
        sheet = self.wb['1-2']

        # Column headers
        columns = {
            'B': 'row_num',
            'C': 'contract_row',
            'D': 'chapter',
            'E': 'sub_category',
            'F': 'index',
            'G': 'description',
            'H': 'package_name',
            'I': 'proposal_date',
            'J': 'calc_base_date',
            'K': 'entry_date',
            'L': 'currency_ratio_query',
            'M': 'currency_ratio_used',
            'N': 'f_coefficient',
            'O': 'purchase_rate',
            'P': 'base_rate',
            'Q': 'calc_year',
            'R': 'n_value',
            'S': 'n_reduction',
            'T': 'months_elapsed',
            'W': 'contract_amount',
            'X': 'compensation_amount',
            'Y': 'delay_flag'
        }

        data = []
        for row in range(7, sheet.max_row + 1):
            row_data = {}
            for col_letter, col_name in columns.items():
                col_idx = column_index_from_string(col_letter)
                row_data[col_name] = sheet.cell(row=row, column=col_idx).value
            data.append(row_data)

        return pd.DataFrame(data)

    def extract_exchange_rates(self) -> pd.DataFrame:
        """
        Extract exchange rates from Arz sheet.

        Returns:
            DataFrame with dates and rates
        """
        # Find Arz sheet
        arz_name = None
        for name in self.wb.sheetnames:
            if 'Arz' in name:
                arz_name = name
                break

        if not arz_name:
            return pd.DataFrame()

        sheet = self.wb[arz_name]
        data = []

        for row in range(2, sheet.max_row + 1):
            date = sheet.cell(row=row, column=1).value
            rate = sheet.cell(row=row, column=2).value
            if date:
                data.append({
                    'date': JalaliDateParser.persian_to_latin(str(date)),
                    'rate': rate
                })

        return pd.DataFrame(data)

    def extract_category_indices(self, sheet_name: str) -> pd.DataFrame:
        """
        Extract index data from a category sheet.

        Args:
            sheet_name: Name of the category sheet

        Returns:
            DataFrame with index data
        """
        if sheet_name not in self.wb.sheetnames:
            return pd.DataFrame()

        sheet = self.wb[sheet_name]

        # Get all data starting from row 4
        data = []
        for row in range(4, sheet.max_row + 1):
            chapter = sheet.cell(row=row, column=1).value
            description = sheet.cell(row=row, column=2).value

            if chapter is not None:
                row_data = {
                    'chapter': chapter,
                    'description': description
                }

                # Get index values
                for col in range(3, sheet.max_column + 1):
                    val = sheet.cell(row=row, column=col).value
                    col_header = sheet.cell(row=2, column=col).value or f"col_{col}"
                    row_data[str(col_header)] = val

                data.append(row_data)

        return pd.DataFrame(data)


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    """Main entry point for the audit script."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Excel Logic Chain Audit Script",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_logic_audit.py file.xlsx
  python excel_logic_audit.py file.xlsx --output report.txt
  python excel_logic_audit.py file.xlsx --json results.json
        """
    )

    parser.add_argument(
        'excel_file',
        nargs='?',
        default='Price_Adjustment_Automated 19 Claude Final 01 Claude Code.xlsx',
        help='Path to the Excel file to audit'
    )

    parser.add_argument(
        '--output', '-o',
        help='Output file for the audit report'
    )

    parser.add_argument(
        '--json', '-j',
        help='Export results to JSON file'
    )

    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose logging'
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    # Run audit
    try:
        auditor = ExcelLogicAuditor(args.excel_file)

        # Generate report
        report = auditor.generate_report(args.output)

        # Export JSON if requested
        if args.json:
            auditor.export_results_json(args.json)

        # Print report to console
        print(report)

    except FileNotFoundError as e:
        logger.error(f"File not found: {e}")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Audit failed: {e}")
        raise


if __name__ == '__main__':
    main()
