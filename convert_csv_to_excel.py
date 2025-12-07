#!/usr/bin/env python3
"""
CSV to Excel Converter for USD to Rials Data
Converts CSV data with mixed date formats into a professionally styled Excel file.
"""

import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def convert_to_persian_numerals(text: str) -> str:
    """Convert English numerals to Persian numerals."""
    persian_numerals = {
        '0': '۰', '1': '۱', '2': '۲', '3': '۳', '4': '۴',
        '5': '۵', '6': '۶', '7': '۷', '8': '۸', '9': '۹'
    }
    result = text
    for eng, per in persian_numerals.items():
        result = result.replace(eng, per)
    return result


def format_gregorian_date(date_str: str) -> str:
    """
    Convert Gregorian date from M/D/YYYY format to YYYY-MM-DD format.
    Handles various input formats.
    """
    try:
        # Try parsing M/D/YYYY format
        dt = datetime.strptime(date_str, "%m/%d/%Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        pass

    try:
        # Try parsing with single digit month/day (e.g., 9/29/1981)
        parts = date_str.split('/')
        if len(parts) == 3:
            month, day, year = parts
            dt = datetime(int(year), int(month), int(day))
            return dt.strftime("%Y-%m-%d")
    except (ValueError, IndexError):
        pass

    # Return original if parsing fails
    return date_str


def format_shamsi_date(date_str: str) -> str:
    """Convert Shamsi date to Persian numerals format."""
    # The date is already in YYYY/MM/DD format, just convert numerals
    return convert_to_persian_numerals(date_str)


def auto_adjust_column_width(worksheet):
    """Auto-adjust column widths based on content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value else ""
                # Account for Persian characters being wider
                cell_length = len(cell_value)
                # Persian text needs more width
                if any('\u0600' <= char <= '\u06FF' or '\u06F0' <= char <= '\u06F9' for char in cell_value):
                    cell_length = int(cell_length * 1.5)
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        # Add some padding
        adjusted_width = max_length + 4
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_styled_excel(input_csv: str, output_xlsx: str):
    """
    Read CSV data and create a professionally styled Excel file.

    Args:
        input_csv: Path to the input CSV file
        output_xlsx: Path for the output Excel file
    """
    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "USD Rates"

    # Define styles
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    cell_alignment_center = Alignment(horizontal="center", vertical="center")
    cell_alignment_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write header row
    headers = ["Gregorian Date", "Shamsi Date", "Source", "USD Rate"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Read and process CSV data
    row_count = 0
    with open(input_csv, 'r', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)

        for row_num, row in enumerate(reader, 2):
            # Extract and format data
            gregorian_date = format_gregorian_date(row['date_gr'])
            shamsi_date = format_shamsi_date(row['date_pr'])
            source = row['source']
            usd_rate = float(row['price_avg']) if row['price_avg'] else 0

            # Write to Excel
            # Gregorian Date
            cell = ws.cell(row=row_num, column=1, value=gregorian_date)
            cell.alignment = cell_alignment_center
            cell.border = thin_border

            # Shamsi Date (Persian numerals)
            cell = ws.cell(row=row_num, column=2, value=shamsi_date)
            cell.alignment = cell_alignment_center
            cell.border = thin_border

            # Source
            cell = ws.cell(row=row_num, column=3, value=source)
            cell.alignment = cell_alignment_center
            cell.border = thin_border

            # USD Rate (numeric)
            cell = ws.cell(row=row_num, column=4, value=usd_rate)
            cell.alignment = cell_alignment_right
            cell.border = thin_border
            cell.number_format = '#,##0'

            row_count += 1

    # Auto-adjust column widths
    auto_adjust_column_width(ws)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Save workbook
    wb.save(output_xlsx)

    return row_count


def main():
    input_file = "/home/user/codex-test/USD2Rials.csv"
    output_file = "/home/user/codex-test/Cleaned_USD_Rates.xlsx"

    print("=" * 60)
    print("CSV to Excel Converter - USD Rates Data")
    print("=" * 60)
    print(f"\nInput file:  {input_file}")
    print(f"Output file: {output_file}")
    print("\nProcessing...")

    try:
        rows_processed = create_styled_excel(input_file, output_file)
        print(f"\nSuccess! Processed {rows_processed:,} data rows.")
        print(f"\nExcel file saved: {output_file}")
        print("\nFeatures applied:")
        print("  - Gregorian dates formatted as YYYY-MM-DD (English numerals)")
        print("  - Shamsi dates converted to Persian numerals")
        print("  - Header row: Bold, centered, light gray background")
        print("  - Column widths auto-adjusted")
        print("  - USD Rate formatted with thousand separators")
        print("  - Header row frozen for easy scrolling")
        print("\nFile is ready for download!")
    except Exception as e:
        print(f"\nError: {e}")
        raise


if __name__ == "__main__":
    main()
