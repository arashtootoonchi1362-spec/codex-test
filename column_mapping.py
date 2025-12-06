"""
Column Mapping Script for Contractor Claims and Employer Approvals

CRITICAL LOGIC:
- Contractor's Claim: Read from contractor.xlsx
  - Look for 'پیمانکار-قطعی' or 'گروه مشارکت' main header -> 'مبلغ' sub-column
- Employer's Approval: Read from employer.xlsx (NOT contractor.xlsx!)
  - Look for 'مشاور زیربنایی' (Infrastructure Consultant) main header -> 'مبلغ (ریال)' sub-column

This script uses DYNAMIC header name matching, not fixed column indices,
to handle variations across different sheets.
"""

import openpyxl
from typing import Tuple, Optional, List, Dict, Any
import re


# Configuration constants
CONTRACTOR_FILE = 'contractor.xlsx'
EMPLOYER_FILE = 'employer.xlsx'

# Main header patterns to search for
CONTRACTOR_CLAIM_HEADERS = ['پیمانکار-قطعی', 'پیمانکار', 'گروه مشارکت']
EMPLOYER_APPROVAL_HEADER = 'مشاور زیربنایی'  # Infrastructure Consultant
EMPLOYER_APPROVAL_HEADER_ALT = 'مشاور زیر بنایی'  # Alternative spelling (with space)

# Sub-header pattern for amount column
AMOUNT_HEADER_PATTERN = 'مبلغ'  # Amount (Rial)

# Sheets to process (common between both files)
DATA_SHEETS = [
    'سازمان دهی و برنامه ریزی پروژه',
    'بیمه نامه ها',
    'تجهیز کارگاه اولیه',
    'تجهیز کارگاه مستمر',
    'خدمات طراحي  مهندسي پیش از اجرا',
    'خدمات طراحي و مهندسي حین اجرا',
    'خدمات مهندسي حين اجرا',
    'خدمات تامین',
    'خدمات اجرایی ',
    'اختتامیه',
]


def find_header_rows(ws, max_search_rows: int = 15) -> Tuple[Optional[int], Optional[int]]:
    """
    Dynamically find the main header row and sub-header row in a worksheet.

    The main header row must contain BOTH:
    - Standard column identifiers like 'ردیف' (Row Number) or 'W.B.S'
    - Section headers like 'گروه مشارکت', 'مشاور زیربنایی', etc.

    This prevents false matches on metadata rows that might contain similar keywords.

    Returns:
        Tuple of (main_header_row, sub_header_row) or (None, None) if not found
    """
    # Standard column identifiers that indicate a proper header row
    standard_columns = ['ردیف', 'W.B.S', 'عنوان فعالیت']
    # Section headers we're looking for
    section_headers = ['گروه مشارکت', 'مشاور زیربنایی', 'مشاور زیر بنایی', 'پیمانکار-قطعی']

    for row_idx in range(1, max_search_rows + 1):
        row_data = [cell.value for cell in ws[row_idx]]
        row_str = ' '.join(str(v) for v in row_data if v)

        # Row must contain BOTH standard column identifiers AND section headers
        has_standard = any(col in row_str for col in standard_columns)
        has_section = any(h in row_str for h in section_headers)

        if has_standard and has_section:
            return row_idx, row_idx + 1

    return None, None


def find_amount_column_for_section(ws, main_header_row: int, sub_header_row: int,
                                    section_headers: List[str]) -> Optional[int]:
    """
    Find the 'مبلغ' (Amount) column index under a specific section header.

    Args:
        ws: Worksheet object
        main_header_row: Row number containing main headers
        sub_header_row: Row number containing sub-headers
        section_headers: List of possible section header names to match

    Returns:
        Column index (1-based) of the 'مبلغ' column, or None if not found
    """
    main_headers = [cell.value for cell in ws[main_header_row]]
    sub_headers = [cell.value for cell in ws[sub_header_row]]

    # Find the section start column
    section_start_col = None
    section_end_col = None

    for col_idx, val in enumerate(main_headers, 1):
        if val and any(h in str(val) for h in section_headers):
            section_start_col = col_idx
            # Find section end (next main header)
            for next_col in range(col_idx + 1, len(main_headers) + 1):
                if main_headers[next_col - 1] is not None:
                    section_end_col = next_col - 1
                    break
            if section_end_col is None:
                section_end_col = len(main_headers)
            break

    if section_start_col is None:
        return None

    # Find 'مبلغ' sub-header within the section range
    for col_idx in range(section_start_col - 1, section_end_col):
        sub_val = sub_headers[col_idx] if col_idx < len(sub_headers) else None
        if sub_val and AMOUNT_HEADER_PATTERN in str(sub_val):
            return col_idx + 1  # Convert to 1-based index

    return None


def get_contractor_claim_column(ws) -> Optional[int]:
    """
    Find the column index for Contractor's Claim amount in contractor.xlsx

    Looks for 'پیمانکار-قطعی' or 'گروه مشارکت' main header -> 'مبلغ' sub-column
    """
    main_header_row, sub_header_row = find_header_rows(ws)
    if main_header_row is None:
        return None

    return find_amount_column_for_section(ws, main_header_row, sub_header_row,
                                          CONTRACTOR_CLAIM_HEADERS)


def get_employer_approval_column(ws) -> Optional[int]:
    """
    Find the column index for Employer's Approval amount in employer.xlsx

    Looks for 'مشاور زیربنایی' (Infrastructure Consultant) main header -> 'مبلغ' sub-column
    """
    main_header_row, sub_header_row = find_header_rows(ws)
    if main_header_row is None:
        return None

    return find_amount_column_for_section(ws, main_header_row, sub_header_row,
                                          [EMPLOYER_APPROVAL_HEADER, EMPLOYER_APPROVAL_HEADER_ALT])


def extract_sheet_data(contractor_wb, employer_wb, sheet_name: str) -> List[Dict[str, Any]]:
    """
    Extract data from a sheet, reading:
    - Contractor's Claim from contractor.xlsx
    - Employer's Approval from employer.xlsx

    Returns:
        List of dictionaries containing row data with contractor claims and employer approvals
    """
    results = []

    # Check if sheet exists in both workbooks
    if sheet_name not in contractor_wb.sheetnames:
        print(f"  Warning: Sheet '{sheet_name}' not found in contractor.xlsx")
        return results
    if sheet_name not in employer_wb.sheetnames:
        print(f"  Warning: Sheet '{sheet_name}' not found in employer.xlsx")
        return results

    contractor_ws = contractor_wb[sheet_name]
    employer_ws = employer_wb[sheet_name]

    # Find header rows for both worksheets
    c_main_row, c_sub_row = find_header_rows(contractor_ws)
    e_main_row, e_sub_row = find_header_rows(employer_ws)

    if c_main_row is None or e_main_row is None:
        print(f"  Warning: Could not find header rows in sheet '{sheet_name}'")
        return results

    # Find the amount columns
    contractor_claim_col = get_contractor_claim_column(contractor_ws)
    employer_approval_col = get_employer_approval_column(employer_ws)

    if contractor_claim_col is None:
        print(f"  Warning: Could not find contractor claim column in sheet '{sheet_name}'")
    if employer_approval_col is None:
        print(f"  Warning: Could not find employer approval column in sheet '{sheet_name}'")

    # Determine data start row (after headers)
    data_start_row = max(c_sub_row, e_sub_row) + 1

    # Read data rows
    # Get WBS column (usually column 2) and Activity Name (column 3) from contractor file
    for row_idx in range(data_start_row, contractor_ws.max_row + 1):
        wbs = contractor_ws.cell(row=row_idx, column=2).value
        activity_name = contractor_ws.cell(row=row_idx, column=3).value

        # Skip empty rows
        if not wbs and not activity_name:
            continue

        contractor_claim = None
        employer_approval = None

        if contractor_claim_col:
            contractor_claim = contractor_ws.cell(row=row_idx, column=contractor_claim_col).value

        if employer_approval_col:
            employer_approval = employer_ws.cell(row=row_idx, column=employer_approval_col).value

        results.append({
            'sheet': sheet_name,
            'row': row_idx,
            'wbs': wbs,
            'activity_name': activity_name,
            'contractor_claim': contractor_claim,
            'employer_approval': employer_approval,
        })

    return results


def process_all_sheets() -> List[Dict[str, Any]]:
    """
    Process all data sheets and extract contractor claims vs employer approvals.

    IMPORTANT:
    - Contractor claims are read from contractor.xlsx
    - Employer approvals are read from employer.xlsx (NOT contractor.xlsx!)
    """
    print("="*70)
    print("COLUMN MAPPING PROCESSOR")
    print("="*70)
    print(f"\nContractor Claims Source: {CONTRACTOR_FILE}")
    print(f"Employer Approvals Source: {EMPLOYER_FILE}")
    print("\nKey Mapping Logic:")
    print("  - Contractor Claim: contractor.xlsx -> 'پیمانکار-قطعی' or 'گروه مشارکت' -> 'مبلغ'")
    print("  - Employer Approval: employer.xlsx -> 'مشاور زیربنایی' -> 'مبلغ (ریال)'")
    print("="*70 + "\n")

    # Load workbooks
    contractor_wb = openpyxl.load_workbook(CONTRACTOR_FILE, read_only=True, data_only=True)
    employer_wb = openpyxl.load_workbook(EMPLOYER_FILE, read_only=True, data_only=True)

    all_results = []

    for sheet_name in DATA_SHEETS:
        print(f"\nProcessing sheet: {sheet_name}")

        sheet_data = extract_sheet_data(contractor_wb, employer_wb, sheet_name)
        all_results.extend(sheet_data)

        # Print summary for this sheet
        if sheet_data:
            total_contractor_claim = sum(
                row['contractor_claim'] for row in sheet_data
                if isinstance(row['contractor_claim'], (int, float))
            )
            total_employer_approval = sum(
                row['employer_approval'] for row in sheet_data
                if isinstance(row['employer_approval'], (int, float))
            )

            print(f"  Rows processed: {len(sheet_data)}")
            print(f"  Total Contractor Claim: {total_contractor_claim:,.0f} Rial")
            print(f"  Total Employer Approval: {total_employer_approval:,.0f} Rial")

    contractor_wb.close()
    employer_wb.close()

    return all_results


def generate_comparison_report(results: List[Dict[str, Any]]) -> None:
    """Generate a summary comparison report."""
    print("\n" + "="*70)
    print("SUMMARY COMPARISON REPORT")
    print("="*70)

    # Group by sheet
    sheets_summary = {}
    for row in results:
        sheet = row['sheet']
        if sheet not in sheets_summary:
            sheets_summary[sheet] = {
                'contractor_total': 0,
                'employer_total': 0,
                'row_count': 0
            }

        if isinstance(row['contractor_claim'], (int, float)):
            sheets_summary[sheet]['contractor_total'] += row['contractor_claim']
        if isinstance(row['employer_approval'], (int, float)):
            sheets_summary[sheet]['employer_total'] += row['employer_approval']
        sheets_summary[sheet]['row_count'] += 1

    grand_contractor = 0
    grand_employer = 0

    print(f"\n{'Sheet Name':<45} {'Contractor Claim':>20} {'Employer Approval':>20} {'Difference':>20}")
    print("-"*105)

    for sheet, summary in sheets_summary.items():
        contractor = summary['contractor_total']
        employer = summary['employer_total']
        diff = contractor - employer

        grand_contractor += contractor
        grand_employer += employer

        print(f"{sheet:<45} {contractor:>20,.0f} {employer:>20,.0f} {diff:>20,.0f}")

    print("-"*105)
    print(f"{'GRAND TOTAL':<45} {grand_contractor:>20,.0f} {grand_employer:>20,.0f} {grand_contractor - grand_employer:>20,.0f}")
    print("="*70)

    if grand_employer == 0 and grand_contractor > 0:
        print("\n*** WARNING: Employer Approval values are ZERO! ***")
        print("This may indicate an issue with column mapping. Please verify:")
        print("  1. employer.xlsx contains 'مشاور زیربنایی' header")
        print("  2. The 'مبلغ' column under this header has data")

    print("\n")


def main():
    """Main entry point."""
    results = process_all_sheets()
    generate_comparison_report(results)
    return results


if __name__ == '__main__':
    main()
