#!/usr/bin/env python3
"""
Script to compare contractor.xlsx and employer.xlsx data and generate disputes report.
Compares data by W.B.S code and identifies differences between the two files.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def get_common_sheets(contractor_xl, employer_xl):
    """Get sheets that exist in both files"""
    contractor_sheets = set(contractor_xl.sheet_names)
    employer_sheets = set(employer_xl.sheet_names)
    return contractor_sheets.intersection(employer_sheets)


def clean_dataframe(df):
    """Clean dataframe by removing fully empty rows and standardizing columns"""
    # Remove rows where all values are NaN
    df = df.dropna(how='all')
    return df


def read_sheet_data(excel_file, sheet_name, header_row=4):
    """Read sheet data with proper header row"""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
        df = clean_dataframe(df)
        return df
    except Exception as e:
        print(f"Error reading {sheet_name}: {e}")
        return None


def find_wbs_column(df):
    """Find the W.B.S column in dataframe"""
    for col in df.columns:
        col_str = str(col).strip()
        if 'W.B.S' in col_str or 'WBS' in col_str.upper():
            return col
    return None


def find_numeric_columns(df):
    """Find columns that contain numeric values for comparison"""
    numeric_cols = []
    for col in df.columns:
        if df[col].dtype in ['float64', 'int64']:
            numeric_cols.append(col)
        else:
            # Check if column has any numeric values
            try:
                numeric_vals = pd.to_numeric(df[col], errors='coerce')
                if numeric_vals.notna().sum() > 0:
                    numeric_cols.append(col)
            except:
                pass
    return numeric_cols


def compare_sheets(contractor_df, employer_df, sheet_name):
    """Compare two dataframes and find disputes"""
    disputes = []

    # Find W.B.S column
    contractor_wbs = find_wbs_column(contractor_df)
    employer_wbs = find_wbs_column(employer_df)

    if contractor_wbs is None or employer_wbs is None:
        # Try to match by row index if no W.B.S column
        return compare_by_index(contractor_df, employer_df, sheet_name)

    # Create lookup dictionaries
    contractor_rows = {}
    for idx, row in contractor_df.iterrows():
        wbs = row[contractor_wbs]
        if pd.notna(wbs):
            contractor_rows[str(wbs).strip()] = row

    employer_rows = {}
    for idx, row in employer_df.iterrows():
        wbs = row[employer_wbs]
        if pd.notna(wbs):
            employer_rows[str(wbs).strip()] = row

    # Find common columns for comparison
    contractor_cols = set(contractor_df.columns)
    employer_cols = set(employer_df.columns)
    common_cols = contractor_cols.intersection(employer_cols)

    # Compare rows
    all_wbs = set(contractor_rows.keys()).union(set(employer_rows.keys()))

    for wbs in all_wbs:
        contractor_row = contractor_rows.get(wbs)
        employer_row = employer_rows.get(wbs)

        if contractor_row is None:
            disputes.append({
                'Sheet': sheet_name,
                'W.B.S': wbs,
                'Column': 'N/A',
                'Contractor Value': 'MISSING',
                'Employer Value': 'EXISTS',
                'Difference': 'Row only in Employer file',
                'Status': 'Dispute'
            })
            continue

        if employer_row is None:
            disputes.append({
                'Sheet': sheet_name,
                'W.B.S': wbs,
                'Column': 'N/A',
                'Contractor Value': 'EXISTS',
                'Employer Value': 'MISSING',
                'Difference': 'Row only in Contractor file',
                'Status': 'Dispute'
            })
            continue

        # Compare values in common columns
        for col in common_cols:
            if col == contractor_wbs:
                continue

            c_val = contractor_row.get(col)
            e_val = employer_row.get(col)

            # Skip if both are NaN
            if pd.isna(c_val) and pd.isna(e_val):
                continue

            # Check for differences
            try:
                # Try numeric comparison
                c_num = float(c_val) if pd.notna(c_val) else 0
                e_num = float(e_val) if pd.notna(e_val) else 0

                if abs(c_num - e_num) > 0.001:  # Tolerance for floating point
                    diff = c_num - e_num
                    disputes.append({
                        'Sheet': sheet_name,
                        'W.B.S': wbs,
                        'Column': str(col),
                        'Contractor Value': c_val,
                        'Employer Value': e_val,
                        'Difference': diff,
                        'Status': 'Dispute'
                    })
            except (ValueError, TypeError):
                # String comparison
                c_str = str(c_val).strip() if pd.notna(c_val) else ''
                e_str = str(e_val).strip() if pd.notna(e_val) else ''

                if c_str != e_str:
                    disputes.append({
                        'Sheet': sheet_name,
                        'W.B.S': wbs,
                        'Column': str(col),
                        'Contractor Value': c_val,
                        'Employer Value': e_val,
                        'Difference': 'Text differs',
                        'Status': 'Dispute'
                    })

    return disputes


def compare_by_index(contractor_df, employer_df, sheet_name):
    """Compare dataframes by row index when W.B.S column is not available"""
    disputes = []

    # Get common columns
    common_cols = set(contractor_df.columns).intersection(set(employer_df.columns))

    max_rows = max(len(contractor_df), len(employer_df))

    for idx in range(max_rows):
        if idx >= len(contractor_df):
            disputes.append({
                'Sheet': sheet_name,
                'W.B.S': f'Row {idx+1}',
                'Column': 'N/A',
                'Contractor Value': 'MISSING',
                'Employer Value': 'EXISTS',
                'Difference': 'Row only in Employer file',
                'Status': 'Dispute'
            })
            continue

        if idx >= len(employer_df):
            disputes.append({
                'Sheet': sheet_name,
                'W.B.S': f'Row {idx+1}',
                'Column': 'N/A',
                'Contractor Value': 'EXISTS',
                'Employer Value': 'MISSING',
                'Difference': 'Row only in Contractor file',
                'Status': 'Dispute'
            })
            continue

        contractor_row = contractor_df.iloc[idx]
        employer_row = employer_df.iloc[idx]

        for col in common_cols:
            c_val = contractor_row.get(col)
            e_val = employer_row.get(col)

            if pd.isna(c_val) and pd.isna(e_val):
                continue

            try:
                c_num = float(c_val) if pd.notna(c_val) else 0
                e_num = float(e_val) if pd.notna(e_val) else 0

                if abs(c_num - e_num) > 0.001:
                    diff = c_num - e_num
                    disputes.append({
                        'Sheet': sheet_name,
                        'W.B.S': f'Row {idx+1}',
                        'Column': str(col),
                        'Contractor Value': c_val,
                        'Employer Value': e_val,
                        'Difference': diff,
                        'Status': 'Dispute'
                    })
            except (ValueError, TypeError):
                c_str = str(c_val).strip() if pd.notna(c_val) else ''
                e_str = str(e_val).strip() if pd.notna(e_val) else ''

                if c_str != e_str:
                    disputes.append({
                        'Sheet': sheet_name,
                        'W.B.S': f'Row {idx+1}',
                        'Column': str(col),
                        'Contractor Value': c_val,
                        'Employer Value': e_val,
                        'Difference': 'Text differs',
                        'Status': 'Dispute'
                    })

    return disputes


def create_disputes_report(all_disputes, output_file):
    """Create formatted Excel report of disputes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Disputes Summary"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    dispute_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ['Sheet', 'W.B.S', 'Column', 'Contractor Value', 'Employer Value', 'Difference', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_idx, dispute in enumerate(all_disputes, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=dispute.get(header, ''))
            cell.border = border
            if dispute.get('Status') == 'Dispute':
                cell.fill = dispute_fill

    # Adjust column widths
    column_widths = [25, 20, 30, 25, 25, 20, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Create summary sheet
    ws_summary = wb.create_sheet("Summary by Sheet")

    # Count disputes by sheet
    sheet_counts = {}
    for dispute in all_disputes:
        sheet = dispute['Sheet']
        sheet_counts[sheet] = sheet_counts.get(sheet, 0) + 1

    # Summary headers
    ws_summary.cell(row=1, column=1, value="Sheet Name").font = header_font
    ws_summary.cell(row=1, column=1).fill = header_fill
    ws_summary.cell(row=1, column=2, value="Number of Disputes").font = header_font
    ws_summary.cell(row=1, column=2).fill = header_fill

    for row_idx, (sheet, count) in enumerate(sorted(sheet_counts.items()), 2):
        ws_summary.cell(row=row_idx, column=1, value=sheet)
        ws_summary.cell(row=row_idx, column=2, value=count)

    # Total row
    total_row = len(sheet_counts) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=2, value=len(all_disputes)).font = Font(bold=True)

    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 20

    wb.save(output_file)
    print(f"Disputes report saved to: {output_file}")
    return len(all_disputes)


def main():
    print("=" * 60)
    print("DISPUTES REPORT GENERATOR")
    print("Comparing contractor.xlsx vs employer.xlsx")
    print("=" * 60)

    # Load Excel files
    print("\nLoading files...")
    contractor_xl = pd.ExcelFile('contractor.xlsx')
    employer_xl = pd.ExcelFile('employer.xlsx')

    print(f"Contractor sheets: {len(contractor_xl.sheet_names)}")
    print(f"Employer sheets: {len(employer_xl.sheet_names)}")

    # Get common sheets
    common_sheets = get_common_sheets(contractor_xl, employer_xl)
    print(f"\nCommon sheets to compare: {len(common_sheets)}")

    all_disputes = []

    # Process each common sheet
    for sheet_name in sorted(common_sheets):
        print(f"\nProcessing sheet: {sheet_name}")

        # Try different header rows (Excel files have varying header positions)
        for header_row in [4, 3, 5, 0]:
            try:
                contractor_df = read_sheet_data(contractor_xl, sheet_name, header_row)
                employer_df = read_sheet_data(employer_xl, sheet_name, header_row)

                if contractor_df is not None and employer_df is not None:
                    if len(contractor_df) > 0 and len(employer_df) > 0:
                        disputes = compare_sheets(contractor_df, employer_df, sheet_name)
                        all_disputes.extend(disputes)
                        print(f"  Found {len(disputes)} disputes (header row: {header_row})")
                        break
            except Exception as e:
                continue

    # Generate report
    print("\n" + "=" * 60)
    print("GENERATING DISPUTES REPORT")
    print("=" * 60)

    total_disputes = create_disputes_report(all_disputes, 'disputes.xlsx')

    print(f"\n{'=' * 60}")
    print(f"SUMMARY")
    print(f"{'=' * 60}")
    print(f"Total disputes found: {total_disputes}")
    print(f"Output file: disputes.xlsx")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
